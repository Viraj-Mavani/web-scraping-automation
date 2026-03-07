import sqlite3
import re
from sqlite3 import Error
import sys
import traceback
import pandas as pd
from bs4 import BeautifulSoup
import time
from selenium import webdriver
import openpyxl
from openpyxl.styles import Font
import chromedriver_autoinstaller
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

BasePath = 'D:\Projects\CedarPython\ADIP-SY603'

File_path= BasePath +'\OP\ADIP-SY603_Output.xlsx'
File_path_txt= BasePath +'\OPtxt\ADIP-SY603_Output.txt'
File_path_count= BasePath +'\Counts\ADIP-SY603_Count.txt'
File_path_error= BasePath +'\Error\ADIP-SY603_Error.xlsx'

persian_alphabet = ["ا", "ب","پ","ت","ث","ج","چ","ح","خ","د","ذ","ر","ز","ژ","س","ش",
                    "ص","ض","ط","ظ","ع","غ","ف","ق","ک","گ","ل","م","ن","و","ه","ی"]


def create_connection(db_file):
    """ create a database connection to the SQLite database
    specified by the db_file
    :param db_file: database file
    :return: Connection object or None
    """
    conn = None
    try:
        conn = sqlite3.connect(db_file)
    except Error as e:
        print(e)
    # except Exception as e:
        # error = traceback.format_exc()
        # print(error)
    return conn


def delete_task(conn, Filepath):
    """
    Delete a task by task id
    :param conn:  Connection to the SQLite database
    :param id: id of the task
    :return:
    """
    sql = 'delete from FileInfoOutput where Filepath=?'
    cur = conn.cursor()
    cur.execute(sql, (Filepath,))
    conn.commit()


def Dereference(obj):
    del obj


def Individual_data(data):
    book1 = openpyxl.load_workbook(File_path)
    sheet1 = book1.active
    global row1

    for item in data:
        Indi_data = ['']*4
        Name_tag = item.find('i', {"class": ['fa', 'fa-user']})
        name = Name_tag.next if Name_tag.next else ''
        Body_data = item.find_all('li')

        phone = ''
        address = ''
        activity = ''

        for li in Body_data:
            if li.next.next == " هاتف : ":
                phone = li.find('span').string if li.find('span') else ''
            elif li.next.next == " العنوان: ":
                address = li.find('span').string if li.find('span') else ''
            else:
                activity = li.string if li.string else ''

        # Create a tuple with all column values
        Indi_data.append(name)
        Indi_data.append(phone)
        Indi_data.append(address)
        Indi_data.append(activity)

        # Check if the row data is already in the set
        # if row_data not in unique_rows:
        sheet1['A{col}'.format(col=row1)] = name
        sheet1['B{col}'.format(col=row1)] = phone
        sheet1['C{col}'.format(col=row1)] = address
        sheet1['D{col}'.format(col=row1)] = activity
        row1 += 1

        try_count=1
        while try_count <= 5:
            try:
                with open(File_path_count,'a') as fh:
                    fh.write('1\n')
                break
            except:
                try_count+=1

        with open(File_path_txt,'a',encoding="utf-8") as fw:
            fw.write("\t".join(map(str,Indi_data))+"\n")

    book1.save(File_path)
    book1.close()


if __name__=='__main__':

    row1=2
    rowError=2

    book1 = openpyxl.Workbook()
    sheet1 = book1.active
    sheet1['A1'] = 'Company Name'
    sheet1['A1'].font = Font(bold=True)
    sheet1['B1'] = 'Phone'
    sheet1['B1'].font = Font(bold=True)
    sheet1['C1'] = 'Address'
    sheet1['C1'].font = Font(bold=True)
    sheet1['D1'] = 'Acitivity'
    sheet1['D1'].font = Font(bold=True)
    book1.save(File_path)
    book1.close()

    bookError = openpyxl.Workbook()
    sheetError = bookError.active
    sheetError['A1'] = 'URL'
    sheetError['A1'].font = Font(bold=True)
    sheetError['B1'] = 'Not Responding'
    sheetError['B1'].font = Font(bold=True)
    sheetError['C1'] = 'Error'
    sheetError['C1'].font = Font(bold=True)
    bookError.save(File_path_error)
    bookError.close()

    Search_headers = ['Company Name','Phone','Adrress','Activity']
    with open(File_path_txt,"w")as f:
        f.write("\t".join(Search_headers)+"\n")
    with open(File_path_count,"w")as f:
        f.write("")

    Base_url = 'http://hamachamber.com/members-index/?ftxt=%D8%A7&searchType=1'
    
    chromedriver_autoinstaller.install()

    options = webdriver.ChromeOptions()
    options.add_argument('--ignore-certificate-errors')
    options.add_argument('--incognito')
    options.add_argument('--disable-logging')
    # options.add_argument('--headless')

    driver = webdriver.Chrome(options=options)

    try:
        for letter in persian_alphabet:
            Base_url_temp = f'http://hamachamber.com/members-index/?ftxt={letter}&ftxt2=&searchType=1&count=1'
            driver.get(Base_url_temp)
            # time.sleep(1)
            
            soup_temp = BeautifulSoup(driver.page_source, 'lxml')
            error_message = soup_temp.find('div', class_='alert alert-danger')
            if error_message:
                break
            last_page_element = soup_temp.find_all('a', class_='page-link')[-2]
            if last_page_element:
                last_page_number = int(last_page_element.get_text(strip=True))
            else:
                last_page_number = 1
            Dereference(soup_temp)
            Dereference(Base_url_temp)
    
            # index = 0
            # while index < last_page_number:
            for index in range(1, last_page_number+1):
                Base_url = f'http://hamachamber.com/members-index/?ftxt={letter}&ftxt2=&searchType=1&count={index}'
                driver.get(Base_url)
                time.sleep(1)
                soup = BeautifulSoup(driver.page_source, 'lxml')
                res = soup.find_all('div', class_='mycalls')[1:]
                Individual_data(res)
                print(f'Success {letter} {index}')

                try:
                    # next_page_locator = (By.XPATH, "//a[@class='page-link'][span[contains(text(), '»')]]")
                    next_page_locator = (By.XPATH, "/html/body/section/div/div/div[1]/div/nav/ul/li[7]/a")
                    next_page = WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable(next_page_locator))
                    next_page.click()
                    # index += 1
                except:
                    pass
                    # break
            print(f'Success {letter}\n\n')
        print(f'Success\n\n')
        
    except:
        error = traceback.format_exc()
        exception_type, exception_object, exception_traceback = sys.exc_info()
        bookError = openpyxl.load_workbook(File_path_error)
        sheetError = bookError.active
        sheetError['A{col}'.format(col=rowError)]= Base_url
        sheetError['B{col}'.format(col=rowError)]= "Not Responding"
        sheetError['C{col}'.format(col=rowError)]= error
        bookError.save(File_path_error)
        bookError.close()
        rowError+=1
    finally:
        data = pd.read_excel(File_path, engine='openpyxl')
        data_file = data.drop_duplicates()
        data_file.to_excel(File_path, index=False)
        exit()

database = r"E:\ADIP Schedulers\NewWorkingDataBase\WorkingDB\InventoryDB.sqldb"

# create a database connection
conn = create_connection(database)
with conn:
    delete_task(conn, File_path)
    delete_task(conn, File_path_txt)
    delete_task(conn, File_path_count)
    delete_task(conn, File_path_error)