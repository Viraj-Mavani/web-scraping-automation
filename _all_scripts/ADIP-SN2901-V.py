import sys
import traceback
import pandas as pd
import sqlite3
import re
from sqlite3 import Error
from bs4 import BeautifulSoup
import time
from selenium import webdriver
import openpyxl
from openpyxl.styles import Font
import chromedriver_autoinstaller
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
# from selenium.webdriver.support.ui import Select
# from selenium.webdriver.chrome.service import Service
# from selenium.webdriver.chrome.options import Options
# from webdriver_manager.chrome import ChromeDriverManager
# import requests

BasePath = 'D:\\Projects\\CedarPython\\ADIP-SN2901'
# BasePath = 'F:\\CedarPython\\ADIP-SN2901'
Listing = {}
Total_URL = 0

File_path_search_OP = BasePath + '\\OP\\ADIP-SN2901_Search_Page.xlsx'
File_path_company_details_OP = BasePath + '\\OP\\ADIP-SN2901_Company_Info.xlsx'


# File_path_search_txt = BasePath + 'OPtxt\\ADIP-SN2901_Search_Page.txt'
# File_path_company_details_txt = BasePath + 'OPtxt\\ADIP-SN2901_Company_Info.txt'
# File_path_search_count = BasePath + 'Counts\\ADIP-SN2901_Count.txt'
File_path_error = BasePath + '\\Error\\ADIP-SN2901_Error.xlsx'


def create_connection(db_file):
    """ create a database connection to the SQLite database
                specified by the db_file
        :param db_file: database file
        :return: Connection object or None
        """
    conn = None
    try:
        conn = sqlite3.connect(db_file)
    except Error as e:
        print(e)

    return conn


def delete_task(conn, Filepath):
    """
        Delete a task by task id
        :param conn:  Connection to the SQLite database
        :param id: id of the task
        :return:
        """
    sql = 'delete from FileInfoOutput where Filepath=?'
    cur = conn.cursor()
    cur.execute(sql, (Filepath,))
    conn.commit()


def Dereference(obj):
    del obj


# def find_element_text(element, selector):
#     try_cnt = 1
#     while try_cnt < 4:
#         try:
#             sub_element = element.find_element(By.CSS_SELECTOR, selector)
#             return sub_element.text.strip()
#         except:
#             try_cnt += 1
#     else:
#         return ''


if __name__ == '__main__':
    File_paths = [File_path_search_OP, File_path_company_details_OP]

    row1 = 2
    book1 = openpyxl.Workbook()
    sheet1 = book1.active
    sheet1['A1'] = "Name"
    sheet1['A1'].font = Font(bold=True)
    sheet1['B1'] = "Year of Creation"
    sheet1['B1'].font = Font(bold=True)
    sheet1['C1'] = "Headquarters"
    sheet1['C1'].font = Font(bold=True)
    sheet1['D1'] = "Legal Form"
    sheet1['D1'].font = Font(bold=True)
    sheet1['E1'] = "Main Activity"
    sheet1['E1'].font = Font(bold=True)
    sheet1['F1'] = "Link"
    sheet1['F1'].font = Font(bold=True)
    book1.save(File_path_search_OP)
    Dereference(sheet1)
    Dereference(book1)

    row2 = 2
    book2 = openpyxl.Workbook()
    sheet2 = book2.active
    sheet2['A1'] = 'Name'
    sheet2['A1'].font = Font(bold=True)
    sheet2['B1'] = 'Siège social'
    sheet2['B1'].font = Font(bold=True)
    sheet2['C1'] = 'Régistre de commerce'
    sheet2['C1'].font = Font(bold=True)
    sheet2['D1'] = 'Ninéa'
    sheet2['D1'].font = Font(bold=True)
    sheet2['E1'] = 'Date Création'
    sheet2['E1'].font = Font(bold=True)
    sheet2['F1'] = 'Localité'
    sheet2['F1'].font = Font(bold=True)
    sheet2['G1'] = 'Gérance'
    sheet2['G1'].font = Font(bold=True)
    sheet2['H1'] = 'Secteur d\'activité'
    sheet2['H1'].font = Font(bold=True)
    sheet2['I1'] = 'Forme Juridique'
    sheet2['I1'].font = Font(bold=True)
    sheet2['J1'] = 'Objet social'
    sheet2['J1'].font = Font(bold=True)
    sheet2['K1'] = 'Exercice social'
    sheet2['K1'].font = Font(bold=True)
    sheet2['L1'] = 'Durée'
    sheet2['L1'].font = Font(bold=True)
    sheet2['M1'] = 'Région'
    sheet2['M1'].font = Font(bold=True)
    sheet2['N1'] = 'Capital'
    sheet2['N1'].font = Font(bold=True)
    sheet2['O1'] = 'Montant des apports en numéraires'
    sheet2['O1'].font = Font(bold=True)
    sheet2['P1'] = 'Description sommaire et l\'évaluation des apports en nature'
    sheet2['P1'].font = Font(bold=True)
    sheet2['Q1'] = 'Nom, prénoms usuels et domicile des associés tenus indéfiniment des dettes sociales'
    sheet2['Q1'].font = Font(bold=True)
    sheet2['R1'] = 'Nom, prénoms et domicile des premiers dirigeants et des premiers commissaires aux comptes'
    sheet2['R1'].font = Font(bold=True)
    sheet2['S1'] = 'Références de l\'immatriculation au registre du commerce et du crédit mobilier'
    sheet2['S1'].font = Font(bold=True)
    sheet2['T1'] = 'Date effective ou prévue du commencement d\'activité'
    sheet2['T1'].font = Font(bold=True)
    sheet2['U1'] = 'Nombre et la valeur nominale des actions souscrites en numéraire'
    sheet2['U1'].font = Font(bold=True)
    sheet2['V1'] = 'Nombre et la valeur nominale des actions attribuées en rémunération de chaque apport en nature'
    sheet2['V1'].font = Font(bold=True)
    sheet2['W1'] = 'Montant de la partie libérée'
    sheet2['W1'].font = Font(bold=True)
    sheet2[
        'X1'] = 'Dispositions statutaires relatives à la constitution des réserves et à la répartition des bénéfices et du boni de liquidation'
    sheet2['X1'].font = Font(bold=True)
    sheet2['Y1'] = 'Avantages particuliers stipulés'
    sheet2['Y1'].font = Font(bold=True)
    sheet2['Z1'] = 'Conditions d\'admission aux assemblées d\'actionnaires et d\'exercice du droit de vote'
    sheet2['Z1'].font = Font(bold=True)
    sheet2['AA1'] = 'Existence de clauses relatives à l\'agrément des cessionnaires d\'actions'
    sheet2['AA1'].font = Font(bold=True)
    sheet2['AB1'] = 'Type Annonces'
    sheet2['AB1'].font = Font(bold=True)
    sheet2['AC1'] = 'Link'
    sheet2['AC1'].font = Font(bold=True)
    book2.save(File_path_company_details_OP)
    Dereference(sheet2)
    Dereference(book2)

    row3 = 2
    book3 = openpyxl.Workbook()
    sheet3 = book3.active
    sheet3['A1'] = 'URL'
    sheet3['A1'].font = Font(bold=True)
    sheet3['B1'] = 'Not Responding'
    sheet3['B1'].font = Font(bold=True)
    sheet3['C1'] = 'Error'
    sheet3['C1'].font = Font(bold=True)
    book3.save(File_path_error)
    book3.close()

    Base_URL = 'https://creationdentreprise.sn/en/finding-business'

    chromedriver_autoinstaller.install()

    options = webdriver.ChromeOptions()
    options.add_argument('--ignore-certificate-errors')
    options.add_argument('--incognito')
    # options.add_argument('--headless')

    driver = webdriver.Chrome(options=options)
    driver.get(Base_URL)
    time.sleep(1)

    try:
        count = 0
        while count < 1747:
            soup = BeautifulSoup(driver.page_source, 'lxml')

            table = soup.find('table', class_="tableheader-processed")
            rows = table.find('tbody').find_all('tr')
            for row in rows:
                cells = row.find_all("td")

                link = row.find('td', class_='views-field-title').a['href']
                title = cells[0].text.strip() if cells[0].text else ''
                date = cells[1].text.strip() if cells[1].text else ''
                siege = cells[2].text.strip() if cells[2].text else ''
                forme_juriduqe = cells[3].text.strip() if cells[3].text else ''
                secteur = cells[4].text.strip() if cells[4].text else ''

                book1 = openpyxl.load_workbook(File_path_search_OP)
                sheet1 = book1.active

                sheet1['A{col}'.format(col=row1)] = title
                sheet1['B{col}'.format(col=row1)] = date
                sheet1['C{col}'.format(col=row1)] = siege
                sheet1['D{col}'.format(col=row1)] = forme_juriduqe
                sheet1['E{col}'.format(col=row1)] = secteur
                sheet1['F{col}'.format(
                    col=row1)] = "https://creationdentreprise.sn" + link

                row1 += 1

                book1.save(File_path_search_OP)
                book1.close()

                # details_link = .find_element(By.TAG_NAME, "a")
                # details_link.click()
                driver.get('https://creationdentreprise.sn' + link)
                # driver.implicitly_wait(3)
                time.sleep(1)

                soup2 = BeautifulSoup(driver.page_source, 'lxml')

                name_element = soup2.find('h1', class_='title-page-societe')
                name = name_element.text.strip() if name_element.text else ''

                divs = soup2.find_all('div', class_="field-items")
                keys = ['hq', 'rc', 'nin', 'date', 'localite', 'gerance', 'secteur', 'forme_juridique', 'objet_social',
                        'exercice_social', 'duree', 'region', 'capital', 'apports', 'description_sommaire', 'nom_associes',
                        'nom_dirigeants', 'references_immatriclation', 'date_commencement', 'valeur_actions',
                        'valeur_nom_actions', 'montant_partie_lib', 'disposition_statutaires', 'avantages_particuliers',
                        'conditions_admission', 'existence_clause', 'type_annonces']
                data_variables = {}
                for index, key in enumerate(keys):
                    if index < len(divs):
                        data_variables[key] = divs[index].text.strip()
                    else:
                        data_variables[key] = ''
                # for index in data_variables.items():
                #     # element = divs[key].find('div', class_="field-item")
                #     # data_variables[key] = divs[key].text.strip() if divs[key].text else ''
                #     data_variables[index] = divs[index].text.strip() if divs[index].text else ''
                # # for index, (key, value) in enumerate(data_variables.items()):
                # #     data_variables[key] = divs[index].text.strip() if index < len(divs) else ''

                book2 = openpyxl.load_workbook(File_path_company_details_OP)
                sheet2 = book2.active
                sheet2['A{col}'.format(col=row2)] = name
                sheet2['B{col}'.format(col=row2)] = data_variables['hq']
                sheet2['C{col}'.format(col=row2)] = data_variables['rc']
                sheet2['D{col}'.format(col=row2)] = data_variables['nin']
                sheet2['E{col}'.format(col=row2)] = data_variables['date']
                sheet2['F{col}'.format(col=row2)] = data_variables['localite']
                sheet2['G{col}'.format(col=row2)] = data_variables['gerance']
                sheet2['H{col}'.format(col=row2)] = data_variables['secteur']
                sheet2['I{col}'.format(
                    col=row2)] = data_variables['forme_juridique']
                sheet2['J{col}'.format(
                    col=row2)] = data_variables['objet_social']
                sheet2['K{col}'.format(
                    col=row2)] = data_variables['exercice_social']
                sheet2['L{col}'.format(col=row2)] = data_variables['duree']
                sheet2['M{col}'.format(col=row2)] = data_variables['region']
                sheet2['N{col}'.format(col=row2)] = data_variables['capital']
                sheet2['O{col}'.format(col=row2)] = data_variables['apports']
                sheet2['P{col}'.format(
                    col=row2)] = data_variables['description_sommaire']
                sheet2['Q{col}'.format(
                    col=row2)] = data_variables['nom_associes']
                sheet2['R{col}'.format(
                    col=row2)] = data_variables['nom_dirigeants']
                sheet2['S{col}'.format(
                    col=row2)] = data_variables['references_immatriclation']
                sheet2['T{col}'.format(
                    col=row2)] = data_variables['date_commencement']
                sheet2['U{col}'.format(
                    col=row2)] = data_variables['valeur_actions']
                sheet2['V{col}'.format(
                    col=row2)] = data_variables['valeur_nom_actions']
                sheet2['W{col}'.format(
                    col=row2)] = data_variables['montant_partie_lib']
                sheet2['X{col}'.format(
                    col=row2)] = data_variables['disposition_statutaires']
                sheet2['Y{col}'.format(
                    col=row2)] = data_variables['avantages_particuliers']
                sheet2['Z{col}'.format(
                    col=row2)] = data_variables['conditions_admission']
                sheet2['AA{col}'.format(
                    col=row2)] = data_variables['existence_clause']
                sheet2['AB{col}'.format(
                    col=row2)] = data_variables['type_annonces']
                sheet2['AC{col}'.format(col=row2)] = link

                row1 += 1

                book2.save(File_path_company_details_OP)
                book2.close()
                print(f"Row Complete {name}")

                driver.back()
            print(f"Page Complete {count+1}")

            next_page_locator = (
                By.XPATH, "//li[@class = 'pager-next last']/a")
            next_page = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable(next_page_locator))
            next_page.click()

            count += 1
            # print("sus {}".format(count))
    except:
        error = traceback.format_exc()
        exception_type, exception_object, exception_traceback = sys.exc_info()
        book3 = openpyxl.load_workbook(File_path_error)
        sheet3 = book3.active
        sheet3['A{col}'.format(col=row3)] = Base_URL
        sheet3['B{col}'.format(col=row3)] = "Not Responding"
        sheet3['C{col}'.format(col=row3)] = error
        book3.save(File_path_error)
        book3.close()
        row3 += 1
    finally:
        data = pd.read_excel(File_path_search_OP, engine='openpyxl')
        data_file = data.drop_duplicates()
        data_file.to_excel(File_path_search_OP, index=False)
        data = pd.read_excel(File_path_company_details_OP, engine='openpyxl')
        data_file = data.drop_duplicates()
        data_file.to_excel(File_path_company_details_OP, index=False)
        exit()
#
# database = r"E:\ADIP Schedulers\NewWorkingDataBase\WorkingDB\InventoryDB.sqldb"
#
# # create a database connection
# conn = create_connection(database)
# with conn:
# 	for File_path in File_paths:
# 		delete_task(conn, File_path)
