import os
import json
import sqlite3
from sqlite3 import Error
from bs4 import BeautifulSoup
import requests
import math
import string
from selenium import webdriver
import openpyxl
from openpyxl.styles import Font
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options

chrome_options = Options()
chrome_options.add_argument('--headless')
chrome_options.add_argument('--no-sandbox')
chrome_options.add_argument('--disable-dev-shm-usage')
chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])

BasePath= 'D:\\Projects\\CedarPython\\ADIP-BW2701'
Total_URL = 0
File_path_General_Details= BasePath + '\\OP\\ADIP-BW2701_General_Details.xlsx'
File_path_Auditors= BasePath + '\\OP\\ADIP-BW2701_Auditors.xlsx'
File_path_Addresses= BasePath + '\\OP\\ADIP-BW2701_Addresses.xlsx'
File_path_Directors= BasePath + '\\OP\\ADIP-BW2701_Directors.xlsx'
File_path_Secretaries= BasePath + '\\OP\\ADIP-BW2701_Secretaries.xlsx'
File_path_Shareholders= BasePath + '\\OP\\ADIP-BW2701_Shareholders.xlsx'
File_path_Share_Allocations= BasePath + '\\OP\\ADIP-BW2701_Share_Allocations.xlsx'
File_path_Proprietors= BasePath + '\\OP\\ADIP-BW2701_Proprietors.xlsx'
File_path_Search_Page_Info= BasePath + '\\OP\\ADIP-BW2701_Search_Page_Info.xlsx'
File_path_General_Details_txt= BasePath + '\\OPtxt\\ADIP-BW2701_General_Details.txt'
File_path_Auditors_txt= BasePath + '\\OPtxt\\ADIP-BW2701_Auditors.txt'
File_path_Addresses_txt= BasePath + '\\OPtxt\\ADIP-BW2701_Addresses.txt'
File_path_Directors_txt= BasePath + '\\OPtxt\\ADIP-BW2701_Directors.txt'
File_path_Secretaries_txt= BasePath + '\\OPtxt\\ADIP-BW2701_Secretaries.txt'
File_path_Shareholders_txt= BasePath + '\\OPtxt\\ADIP-BW2701_Shareholders.txt'
File_path_Share_Allocations_txt= BasePath + '\\OPtxt\\ADIP-BW2701_Share_Allocations.txt'
File_path_Proprietors_txt= BasePath + '\\OPtxt\\ADIP-BW2701_Proprietors.txt'
File_path_Search_Page_Info_txt= BasePath + '\\OPtxt\\ADIP-BW2701_Search_Page_Info.txt'
File_path_search_count= BasePath + '\\Counts\\ADIP-BW2701_Count.txt'
Error_File= BasePath + '\\Error\\ADIP-BW2701_Error.xlsx'

alphabet = list(string.ascii_lowercase)

def create_connection(db_file):
	""" create a database connection to the SQLite database
		specified by the db_file
	:param db_file: database file
	:return: Connection object or None
	"""
	conn = None
	try:
		conn = sqlite3.connect(db_file)
	except Error as e:
		print(e)

	return conn

def delete_task(conn, Filepath):
	"""
	Delete a task by task id
	:param conn:  Connection to the SQLite database
	:param id: id of the task
	:return:
	"""
	sql = 'delete from FileInfoOutput where Filepath=?'
	cur = conn.cursor()
	cur.execute(sql, (Filepath,))
	conn.commit()
	
def Dereference(obj):
	del obj

def General_Details(ID,data,row,isCompany):
	book2 = openpyxl.load_workbook(File_path_General_Details)
	sheet2 = book2.active
	General_Details_data = ['']*21
	Main_data = data[data[data[ID]['children'][0]]['children'][0]]['children']
	for i in Main_data:
		if 'attribute' in data[i] and data[i]['attribute'] == '_businessIdentifier':
			sheet2['J{col}'.format(col=row)] = data[i]['attributeValue'] if 'attributeValue' in data[i] else ''
			General_Details_data[9] = data[i]['attributeValue'] if 'attributeValue' in data[i] else ''
		elif 'attribute' in data[i] and data[i]['attribute'] == 'ForeignCompanyDocumentLodgedYn':
			sheet2['O{col}'.format(col=row)] = data[i]['attributeValue'] if 'attributeValue' in data[i] else ''
			General_Details_data[14] = data[i]['attributeValue'] if 'attributeValue' in data[i] else ''
		elif 'attribute' in data[i] and data[i]['attribute'] == 'Exempt':
			sheet2['P{col}'.format(col=row)] = data[i]['text']['dcText'] if ('text' in data[i] and 'dcText' in data[i]['text']) else ''
			General_Details_data[15] = data[i]['text']['dcText'] if ('text' in data[i] and 'dcText' in data[i]['text']) else ''
		elif 'attribute' in data[i] and data[i]['attribute'] == 'RegistrationDate':
			if 'text' in data[i] and 'label' in data[i]['text'] and data[i]['text']['label'] == 'Incorporation Date':
				sheet2['Q{col}'.format(col=row)] = data[i]['attributeValue'] if 'attributeValue' in data[i] else ''
				General_Details_data[16] = data[i]['attributeValue'] if 'attributeValue' in data[i] else ''
			if 'text' in data[i] and 'label' in data[i]['text'] and data[i]['text']['label'] == 'Registration Date':
				sheet2['E{col}'.format(col=row)] = data[i]['attributeValue'] if 'attributeValue' in data[i] else ''
				General_Details_data[4] = data[i]['attributeValue'] if 'attributeValue' in data[i] else ''			
		elif 'attribute' in data[i] and data[i]['attribute'] == 'ReregistrationDate':
			sheet2['R{col}'.format(col=row)] = data[i]['attributeValue'] if 'attributeValue' in data[i] else ''
			General_Details_data[17] = data[i]['attributeValue'] if 'attributeValue' in data[i] else ''
		elif 'attribute' in data[i] and data[i]['attribute'] == 'OwnConstitutionYn':
			sheet2['S{col}'.format(col=row)] = data[i]['text']['dcText'] if ('text' in data[i] and 'dcText' in data[i]['text']) else ''
			General_Details_data[18] = data[i]['text']['dcText'] if ('text' in data[i] and 'dcText' in data[i]['text']) else ''
		elif 'attribute' in data[i] and data[i]['attribute'] == 'FilingMonth':
			sheet2['T{col}'.format(col=row)] = data[i]['text']['dcText'] if ('text' in data[i] and 'dcText' in data[i]['text']) else ''
			General_Details_data[19] = data[i]['text']['dcText'] if ('text' in data[i] and 'dcText' in data[i]['text']) else ''
		elif 'attribute' in data[i] and data[i]['attribute'] == 'ARLastFiledDate':
			sheet2['U{col}'.format(col=row)] = data[i]['attributeValue'] if 'attributeValue' in data[i] else ''
			General_Details_data[20] = data[i]['attributeValue'] if 'attributeValue' in data[i] else ''
		elif 'text' in data[i] and 'label' in data[i]['text'] and data[i]['text']['label'] == 'Status History':
			Status_All = data[data[i]['children'][0]]['text']['row']
			Status = Status_All.split('(')[0].strip()
			Dates = Status_All.split('(')[1]
			Start = Dates.split('to')[0].strip()
			End = Dates.split('to')[1].strip()
			sheet2['L{col}'.format(col=row)] = Status
			General_Details_data[11] = Status
			sheet2['M{col}'.format(col=row)] = Start
			General_Details_data[12] = Start
			sheet2['N{col}'.format(col=row)] = End[:-1]
			General_Details_data[13] = End[:-1]
		elif 'children' in data[i]:
			if 'domain' in data[data[i]['children'][0]] and data[data[i]['children'][0]]['domain'] == 'Status':
				if 'text' in data[data[i]['children'][0]] and 'row' in data[data[i]['children'][0]]['text']:
					if isCompany == 1:
						sheet2['K{col}'.format(col=row)] = data[data[i]['children'][0]]['text']['row']
						General_Details_data[10] = data[data[i]['children'][0]]['text']['row']
					else:
						sheet2['D{col}'.format(col=row)] = data[data[i]['children'][0]]['text']['row']
						General_Details_data[3] = data[data[i]['children'][0]]['text']['row']
		elif 'attribute' in data[i] and data[i]['attribute'] == 'RenewalMonth':
			sheet2['I{col}'.format(col=row)] = data[i]['text']['dcText'] if ('text' in data[i] and 'dcText' in data[i]['text']) else ''
			General_Details_data[8] = data[i]['text']['dcText'] if ('text' in data[i] and 'dcText' in data[i]['text']) else ''
	for j in data:
		if 'DC_' in j:
			break
		else:
			if 'attribute' in data[j] and data[j]['attribute'] == '_ClassificationDescription':
				sheet2['G{col}'.format(col=row)] = data[j]['attributeValue'] if 'attributeValue' in data[j] else ''
				General_Details_data[6] = data[j]['attributeValue'] if 'attributeValue' in data[j] else ''
			elif 'attribute' in data[j] and data[j]['attribute'] == 'CommencementDate':
				sheet2['H{col}'.format(col=row)] = data[j]['attributeValue'] if 'attributeValue' in data[j] else ''
				General_Details_data[7] = data[j]['attributeValue'] if 'attributeValue' in data[j] else ''
			elif 'widget' in data[j] and data[j]['widget'] == 'attribute-value-list':
				if 'text' in data[j] and 'label' in data[j]['text'] and data[j]['text']['label'] == 'Name History':
					Name_History = data[data[j]['children'][0]]['text']['row']
					Name = ' '.join(Name_History.split(' ')[:-3])
					Date = Name_History.split(' ')
					StartDate = Date[-3]
					EndDate = Date[-1]
					sheet2['A{col}'.format(col=row)] = Name
					sheet2['B{col}'.format(col=row)] = StartDate[1:]
					sheet2['C{col}'.format(col=row)] = EndDate[:-1]
					General_Details_data[0] = Name
					General_Details_data[1] = StartDate
					General_Details_data[2] = EndDate[:-1]
	with open(File_path_General_Details_txt,"a", encoding='utf-8')as f:
		f.write("\t".join(map(str,General_Details_data))+"\n")
	book2.save(File_path_General_Details)
	book2.close()
	Dereference(sheet2)
	Dereference(book2)
	row+=1
	return row 

def Addresses(data,UIN):
	global row3
	book3 = openpyxl.load_workbook(File_path_Addresses)
	sheet3 = book3.active
	Addresses_Data = ['']*7
	Addresses_Data[6] = UIN
	sheet3['G{col}'.format(col=row3)] = UIN
	for i in data:
		if 'DC_' in i:
			break
		else:
			if 'domain' in data[i] and data[i]['domain'] == 'RegisteredOfficeAddress':
				if 'text' in data[i] and 'singleline' in data[i]['text']:
					sheet3['A{col}'.format(col=row3)] = data[i]['text']['row'].strip()
					Addresses_Data[0] = data[i]['text']['row'].strip()
			elif 'domain' in data[i] and data[i]['domain'] == 'EntityPostalAddress':
				if 'text' in data[i] and 'row' in data[i]['text']:
					Postal = data[i]['text']['row']
					if '(' in Postal:
						sheet3['D{col}'.format(col=row3)] = Postal.strip()
						Addresses_Data[3] = Postal.strip()
					else:
						sheet3['C{col}'.format(col=row3)] = Postal.strip()
						Addresses_Data[2] = Postal.strip()
			elif 'domain' in data[i] and data[i]['domain'] == 'PrincipalPlaceOfBusinessAddress':
				if 'text' in data[i] and 'singleline' in data[i]['text']:
					sheet3['E{col}'.format(col=row3)] = data[i]['text']['row'].strip()
					Addresses_Data[4] = data[i]['text']['row'].strip()
			elif 'widget' in data[i] and data[i]['widget'] == 'attribute-value-list':
				if 'text' in data[i] and 'label' in data[i]['text'] and data[i]['text']['label'] == 'Previous Principal Places of Business':
					Prev_postal = data[data[i]['children'][0]]['text']['row']
					Prev_postal = Prev_postal.split(' ')
					sheet3['F{col}'.format(col=row3)] = ' '.join(Prev_postal[:-3])
					Addresses_Data[5] = ' '.join(Prev_postal[:-3])
				elif 'text' in data[i] and 'label' in data[i]['text'] and data[i]['text']['label'] == 'Previous Registered Office Addresses':
					Prev = data[data[i]['children'][0]]['text']['row']
					Prev = Prev.split(' ')
					sheet3['B{col}'.format(col=row3)] = ' '.join(Prev[:-3])
					Addresses_Data[1] = ' '.join(Prev[:-3])
			# elif 'nodetype' in data[i] and data[i]['nodetype'] == 'text':
			# 	if 'dos' in data[i] and data[i]['dos'][0] == 'css-header-name':
			# 		sheet3['F{col}'.format(col=row3)] = data[i]['text']['singleline'].strip() if ('text' in data[i] and 'singleline' in data[i]['text']) else ''
	with open(File_path_Addresses_txt,"a", encoding='utf-8')as f:
		f.write("\t".join(map(str,Addresses_Data))+"\n")
	book3.save(File_path_Addresses)
	book3.close()
	Dereference(sheet3)
	Dereference(book3)
	row3+=1

def Directors(data,UIN):
	book4 = openpyxl.load_workbook(File_path_Directors)
	sheet4 = book4.active
	Directors_Data = ['']*7
	global row4
	collections = []
	for i in data:
		if 'DC_' in i:
			break
		else: 
			if 'domain' in data[i] and data[i]['domain'] == 'IndividualDirector':
				collections.append(i)
	for id in collections:
		sheet4['G{col}'.format(col=row4)] = UIN
		Directors_Data[6] = UIN
		if 'text' in data[id] and 'singleline' in data[id]['text']:
			sheet4['A{col}'.format(col=row4)] = data[id]['text']['singleline']
			Directors_Data[0] = data[id]['text']['singleline']
		Detail_ID = data[id]['children'][1]
		Nationality_ID = data[Detail_ID]['children'][0]
		Nationality_ID = data[Nationality_ID]['children'][1]
		if 'attribute' in data[Nationality_ID] and 'text' in data[Nationality_ID] and data[Nationality_ID]['attribute'] == 'Nationality':
			sheet4['B{col}'.format(col=row4)] = data[Nationality_ID]['text']['dcText']
			Directors_Data[1] = data[Nationality_ID]['text']['dcText']
		Resident = data[Detail_ID]['children'][2]
		Resident = data[Resident]['children'][0]
		Resident = data[Resident]['children'][0]
		if 'domain' in data[Resident] and 'text' in data[Resident] and data[Resident]['domain'] == 'ResidentialAddress':
			sheet4['C{col}'.format(col=row4)] = data[Resident]['text']['row'].strip()
			Directors_Data[2] = data[Resident]['text']['row'].strip()
		Postal = data[Detail_ID]['children'][3]
		Postal = data[Postal]['children'][0]
		Postal = data[Postal]['children'][0]
		if 'domain' in data[Postal] and 'text' in data[Postal] and data[Postal]['domain'] == 'ServiceAddress':
			sheet4['D{col}'.format(col=row4)] = data[Postal]['text']['row'].strip()
			Directors_Data[3] = data[Postal]['text']['row'].strip()
		Appointment = data[Detail_ID]['children'][4]
		Appointment = data[Appointment]['children']
		for ids in Appointment:
			if 'children' in data[ids]:
				for j in range(0,len(data[ids]['children'])):
					Date_ID = data[ids]['children'][j]
					if 'attribute' in data[Date_ID] and 'attributeValue' in data[Date_ID]:
						if data[Date_ID]['attribute'] == 'StartDate':
							sheet4['E{col}'.format(col=row4)] = data[Date_ID]['attributeValue'].strip()
							Directors_Data[4] = data[Date_ID]['attributeValue'].strip()
						elif data[Date_ID]['attribute'] == 'EndDate':
							sheet4['F{col}'.format(col=row4)] = data[Date_ID]['attributeValue'].strip()
							Directors_Data[5] = data[Date_ID]['attributeValue'].strip()
		
		row4+=1
	with open(File_path_Directors_txt,"a", encoding='utf-8')as f:
		f.write("\t".join(map(str,Directors_Data))+"\n")
	book4.save(File_path_Directors)
	book4.close()
	Dereference(sheet4)
	Dereference(book4)

def Secretaries(data,UIN):
	book6 = openpyxl.load_workbook(File_path_Secretaries)
	sheet6 = book6.active
	global row6
	collection = {}
	isCompany = 0
	for i in data:
		if 'DC_' in i:
			break
		else:
			if 'domain' in data[i] and data[i]['domain'] == 'EntitySecretary':
				isCompany = 1
				collection[i] = isCompany
			elif 'domain' in data[i] and data[i]['domain'] == 'IndividualSecretary':
				isCompany = 0
				collection[i] = isCompany
	for key,value in collection.items():
			Secretaries_Data = ['']*14
			sheet6['M{col}'.format(col=row6)] = UIN
			Secretaries_Data[13] = UIN
			if value == 1:
				Details_ID = data[key]['children'][1]
				Secretary = data[Details_ID]['children'][0]
				Secretary = data[Secretary]['children'][0]
				for id in data[Secretary]['children']:
					if 'attribute' in data[id] and 'attributeValue' in data[id] and data[id]['attribute'] == 'Name':
						sheet6['A{col}'.format(col=row6)] = data[id]['attributeValue']
						Secretaries_Data[0] = data[id]['attributeValue']
					elif 'attribute' in data[id] and 'attributeValue' in data[id] and data[id]['attribute'] == 'EntityNumber':
						sheet6['B{col}'.format(col=row6)] = data[id]['attributeValue']
						Secretaries_Data[1] = data[id]['attributeValue']
					elif 'text' in data[id] and 'label' in data[id]['text'] and data[id]['text']['label'] == 'Registered Office Address':
						ad = data[id]['children'][0]
						sheet6['C{col}'.format(col=row6)] = data[ad]['text']['row'].strip()
						Secretaries_Data[2] = data[ad]['text']['row'].strip()
				Representative = data[Details_ID]['children'][1]
				for j in data[Representative]['children']:
					if 'attribute' in data[j] and 'attributeValue' in data[j] and data[j]['attribute'] == 'Name':
						sheet6['D{col}'.format(col=row6)] = data[j]['attributeValue'].strip()
						Secretaries_Data[3] = data[j]['attributeValue'].strip()
					elif 'domain' in data[j] and 'text' in data[j] and data[j]['domain'] == 'ServiceAddress':
						sheet6['E{col}'.format(col=row6)] = data[j]['text']['row'].strip()
						Secretaries_Data[4] = data[j]['text']['row'].strip()
				Additional_Details = data[Details_ID]['children'][2]
				for k in data[Additional_Details]['children']:
					if 'children' in data[k]:
						for l in data[k]['children']:
							if data[l]['attribute'] == 'StartDate':
								sheet6['F{col}'.format(col=row6)] = data[l]['attributeValue'].strip()
								Secretaries_Data[5] = data[l]['attributeValue'].strip()
				sheet6['L{col}'.format(col=row6)] = isCompany
				Secretaries_Data[11] = isCompany
				row6+=1
				with open(File_path_Secretaries_txt,"a", encoding='utf-8')as f:
					f.write("\t".join(map(str,Secretaries_Data))+"\n")
			else:
				sheet6['G{col}'.format(col=row6)] = data[key]['text']['row'].strip()
				Secretaries_Data[6] = data[key]['text']['row'].strip()
				Details = data[key]['children'][1]
				Nationality = data[Details]['children'][0]
				Nationality = data[Nationality]['children'][1]
				if 'text' in data[Nationality]:
					sheet6['H{col}'.format(col=row6)] = data[Nationality]['text']['dcText'].strip()
					Secretaries_Data[7] = data[Nationality]['text']['dcText'].strip()
				Address = data[Details]['children'][2]
				Address = data[data[Address]['children'][0]]['children'][0]
				sheet6['I{col}'.format(col=row6)] = data[Address]['text']['row'].strip()
				Secretaries_Data[8] = data[Address]['text']['row'].strip()
				Postal = data[Details]['children'][3]
				Postal = data[data[Postal]['children'][0]]['children'][0]
				sheet6['J{col}'.format(col=row6)] = data[Postal]['text']['row'].strip()
				Secretaries_Data[9] = data[Postal]['text']['row'].strip()
				Additional = data[Details]['children'][4]
				for a in data[Additional]['children']:
					if 'children' in data[a]:
						for b in data[a]['children']:
							if 'attribute' in data[b] and data[b]['attribute'] == 'StartDate':
								sheet6['K{col}'.format(col=row6)] = data[b]['attributeValue'].strip()
								Secretaries_Data[10] = data[b]['attributeValue'].strip()
							elif 'attribute' in data[b] and data[b]['attribute'] == 'EndDate':
								sheet6['N{col}'.format(col=row6)] = data[b]['attributeValue'].strip()
								Secretaries_Data[12] = data[b]['attributeValue'].strip()
				sheet6['L{col}'.format(col=row6)] = isCompany
				Secretaries_Data[11] = isCompany
				row6+=1
				with open(File_path_Secretaries_txt,"a", encoding='utf-8')as f:
					f.write("\t".join(map(str,Secretaries_Data))+"\n")
	book6.save(File_path_Secretaries)
	Dereference(sheet6)
	Dereference(book6)

def Auditors(data,UIN):
	book5 = openpyxl.load_workbook(File_path_Auditors)
	sheet5 = book5.active
	Auditors_Data = ['']*6
	global row5
	collections = []
	for i in data:
		if 'DC_' in i:
			break
		else: 
			if 'domain' in data[i] and data[i]['domain'] == 'IndividualAuditor':
				collections.append(i)

	for id in collections:
		sheet5['F{col}'.format(col=row5)] = UIN
		Auditors_Data[5] = UIN
		if 'text' in data[id] and 'singleline' in data[id]['text']:
			sheet5['A{col}'.format(col=row5)] = data[id]['text']['singleline']
			Auditors_Data[0] = data[id]['text']['singleline']
		Detail_ID = data[id]['children'][1]
		Nationality_ID = data[Detail_ID]['children'][0]
		Nationality_ID = data[Nationality_ID]['children'][1]
		if 'attribute' in data[Nationality_ID] and 'text' in data[Nationality_ID] and data[Nationality_ID]['attribute'] == 'Nationality':
			sheet5['B{col}'.format(col=row5)] = data[Nationality_ID]['text']['dcText']
			Auditors_Data[1] = data[Nationality_ID]['text']['dcText']
		Resident = data[Detail_ID]['children'][2]
		Resident = data[Resident]['children'][0]
		Resident = data[Resident]['children'][0]
		if 'domain' in data[Resident] and 'text' in data[Resident] and data[Resident]['domain'] == 'ResidentialAddress':
			sheet5['C{col}'.format(col=row5)] = data[Resident]['text']['row'].strip()
			Auditors_Data[2] = data[Resident]['text']['row'].strip()
		Appointment = data[Detail_ID]['children'][3]
		Appointment = data[Appointment]['children']
		for ids in Appointment:
			if 'children' in data[ids]:
				for j in range(0,len(data[ids]['children'])):
					Date_ID = data[ids]['children'][j]
					if 'attribute' in data[Date_ID] and 'attributeValue' in data[Date_ID]:
						if data[Date_ID]['attribute'] == 'StartDate':
							sheet5['D{col}'.format(col=row5)] = data[Date_ID]['attributeValue'].strip()
							Auditors_Data[3] = data[Date_ID]['attributeValue'].strip()
						elif data[Date_ID]['attribute'] == 'EndDate':
							sheet5['E{col}'.format(col=row5)] = data[Date_ID]['attributeValue'].strip()
							Auditors_Data[4] = data[Date_ID]['attributeValue'].strip()
		
		row5+=1
		with open(File_path_Auditors_txt,"a", encoding='utf-8')as f:
			f.write("\t".join(map(str,Auditors_Data))+"\n")
	book5.save(File_path_Auditors)
	book5.close()
	Dereference(sheet5)
	Dereference(book5)

def Shareholders(data,UIN):
	book7 = openpyxl.load_workbook(File_path_Shareholders)
	sheet7 = book7.active
	global row7
	collection = {}
	isCompany = 0
	for ids in data:
		if 'DC_' in ids:
			break
		else:
			if 'domain' in data[ids] and data[ids]['domain'] == 'IndividualShareholder':
				isCompany = 0
				collection[ids] = isCompany
			elif 'domain' in data[ids] and data[ids]['domain'] == 'EntityShareholder':
				isCompany = 1
				collection[ids] = isCompany
			elif 'domain' in data[ids] and data[ids]['domain'] == 'OtherShareholder':
				isCompany = 2
				collection[ids] = isCompany
	for key,value in collection.items():
		Shareholders_Data = ['']*18
		sheet7['R{col}'.format(col=row7)] = UIN
		Shareholders_Data[17] = UIN
		if value == 0:
			sheet7['K{col}'.format(col=row7)] = data[key]['text']['singleline'].strip()
			Shareholders_Data[10] = data[key]['text']['singleline'].strip()
			Details = data[key]['children'][1]
			Nationality = data[Details]['children'][0]
			Nationality = data[Nationality]['children'][1]
			if 'text' in data[Nationality]:
				sheet7['L{col}'.format(col=row7)] = data[Nationality]['text']['dcText']
				Shareholders_Data[11] = data[Nationality]['text']['dcText']
			Address = data[Details]['children'][2]
			Address = data[data[Address]['children'][0]]['children'][0]
			sheet7['M{col}'.format(col=row7)] = data[Address]['text']['row'].strip()
			Shareholders_Data[12] = data[Address]['text']['row'].strip()
			Postal = data[Details]['children'][3]
			Postal = data[data[Postal]['children'][0]]['children'][0]
			sheet7['N{col}'.format(col=row7)] = data[Postal]['text']['row'].strip()
			Shareholders_Data[13] = data[Postal]['text']['row'].strip()
			Nominee = data[Details]['children'][4]
			Nominee = data[Nominee]['children'][0]
			sheet7['O{col}'.format(col=row7)] = data[Nominee]['text']['dcText'].strip()
			Shareholders_Data[14] = data[Nominee]['text']['dcText'].strip()
			Additional = data[Details]['children'][5]
			for i in data[Additional]['children']:
				if 'children' in data[i]:
					for m in data[i]['children']:
						if 'attribute' in data[m] and data[m]['attribute'] == 'StartDate':
							sheet7['P{col}'.format(col=row7)] = data[m]['attributeValue']
							Shareholders_Data[15] = data[m]['attributeValue']
			sheet7['Q{col}'.format(col=row7)] = value
			Shareholders_Data[16] = value
			row7+=1
			with open(File_path_Shareholders_txt,"a", encoding='utf-8')as f:
				f.write("\t".join(map(str,Shareholders_Data))+"\n")
		elif value == 1:
			Details = data[key]['children'][1]
			Share = data[Details]['children'][0]
			for j in data[Share]['children']:
				if 'attribute' in data[j] and data[j]['attribute'] == 'Name':
					sheet7['A{col}'.format(col=row7)] = data[j]['attributeValue'].strip()
					Shareholders_Data[0] = data[j]['attributeValue'].strip()
				elif 'attribute' in data[j] and data[j]['attribute'] == 'EntityNumber':
					if 'attributeValue' in data[j]:
						sheet7['B{col}'.format(col=row7)] = data[j]['attributeValue'].strip()
						Shareholders_Data[1] = data[j]['attributeValue'].strip()
				elif 'text' in data[j] and 'label' in data[j]['text'] == 'Registered Office Address':
					Ad = data[j]['children'][0]
					sheet7['C{col}'.format(col=row7)] = data[Ad]['text']['row'].strip()
					Shareholders_Data[2] = data[Ad]['text']['row'].strip()
			Postal = data[Details]['children'][1]
			Postal = data[Postal]['children'][0]
			sheet7['D{col}'.format(col=row7)] = data[Postal]['text']['row'].strip()
			Shareholders_Data[3] = data[Postal]['text']['row'].strip()
			Additional_Details = data[Details]['children'][3]
			for k in data[Additional_Details]['children']:
				if 'attribute' in data[k] and data[k]['attribute'] == 'NomineeYn':
					sheet7['E{col}'.format(col=row7)] = data[k]['text']['dcText'].strip() if 'dcText' in data[k]['text'] else ''
					Shareholders_Data[4] = data[k]['text']['dcText'].strip() if 'dcText' in data[k]['text'] else ''
				elif 'children' in data[k]:
					for l in data[k]['children']:
						if 'attribute' in data and data[l]['attribute'] == 'StartDate':
							sheet7['F{col}'.format(col=row7)] = data[l]['attributeValue'].strip()
							Shareholders_Data[5] = data[l]['attributeValue'].strip()
			sheet7['Q{col}'.format(col=row7)] = value
			Shareholders_Data[16] = value
			row7+=1
			with open(File_path_Shareholders_txt,"a", encoding='utf-8')as f:
				f.write("\t".join(map(str,Shareholders_Data))+"\n")
		elif value == 2:
			Details = data[key]['children'][1]
			Shareholders_Details = data[Details]['children'][0]
			for i in data[Shareholders_Details]['children']:
				if 'attribute' in data[i] and data[i]['attribute'] == 'Name':
					sheet7['G{col}'.format(col=row7)] = data[i]['attributeValue'].strip()
					Shareholders_Data[6] = data[i]['attributeValue'].strip()
				elif 'text' in data[i] and data[i]['text']['label'] == 'Address':
					sheet7['H{col}'.format(col=row7)] = data[data[i]['children'][0]]['text']['singleline'].strip()
					Shareholders_Data[7] = data[data[i]['children'][0]]['text']['singleline'].strip()
				elif 'text' in data[i] and data[i]['text']['label'] == 'Postal Address':
					sheet7['N{col}'.format(col=row7)] = data[data[i]['children'][0]]['text']['singleline'].strip()
					Shareholders_Data[13] = data[data[i]['children'][0]]['text']['singleline'].strip()
				elif 'attribute' in data[i] and data[i]['attribute'] == 'EntityNumber':
					sheet7['I{col}'.format(col=row7)] = data[i]['attributeValue'].strip() if 'attributeValue' in data[i] else ''
					Shareholders_Data[8] = data[i]['attributeValue'].strip() if 'attributeValue' in data[i] else ''
				elif 'attribute' in data[i] and data[i]['attribute'] == 'CountryOfOrigin':
					sheet7['J{col}'.format(col=row7)] = data[i]['text']['dcText'].strip() if 'dcText' in data[i]['text'] else ''
					Shareholders_Data[9] = data[i]['text']['dcText'].strip() if 'dcText' in data[i]['text'] else ''
			Nominee = data[Details]['children'][2]
			if 'text' in data[Nominee] and data[Nominee]['text']['label'] == 'Nominee and Beneficial Owner Details':
					sheet7['O{col}'.format(col=row7)] = data[data[Nominee]['children'][0]]['text']['dcText'].strip()
					Shareholders_Data[14] = data[data[Nominee]['children'][0]]['text']['dcText'].strip()
			Appointment = data[Details]['children'][3]
			for m in data[Appointment]['children']:
				if 'children' in data[m]:
					for j in data[m]['children']:
						if 'attribute' in data[j] and data[j]['attribute'] == 'StartDate':
							sheet7['P{col}'.format(col=row7)] = data[j]['attributeValue'].strip()
							Shareholders_Data[15] = data[j]['attributeValue'].strip()
			sheet7['Q{col}'.format(col=row7)] = value
			Shareholders_Data[16] = value
			row7+=1
			with open(File_path_Shareholders_txt,"a", encoding='utf-8')as f:
				f.write("\t".join(map(str,Shareholders_Data))+"\n")

	book7.save(File_path_Shareholders)
	book7.close()
	Dereference(sheet7)
	Dereference(book7)

def Share_Allocations(data,UIN):
	book8 = openpyxl.load_workbook(File_path_Share_Allocations)
	sheet8 = book8.active
	global row8
	Share_Allocations_Data = ['']*5
	Type = ''
	Total_Shares = ''
	for i in data:
		if 'DC_' in i:
			break
		else:
			if 'attribute' in data[i] and data[i]['attribute'] == 'ShareAllocationType':
				Type += data[i]['text']['dcText'].strip() if 'dcText' in data[i]['text'] else ''
			elif 'attribute' in data[i] and data[i]['attribute'] == 'TotalShares':
				Total_Shares += data[i]['attributeValue'].strip()
	for j in data:
		if 'DC_' in j:
			break
		else:
			if 'domain' in data[j] and data[j]['domain'] == 'OwnershipBundle':
				sheet8['E{col}'.format(col=row8)] = UIN
				Share_Allocations_Data[4] = UIN
				Shares = data[j]['children'][0]
				sheet8['B{col}'.format(col=row8)] = data[Shares]['attributeValue'].strip()
				Share_Allocations_Data[1] = data[Shares]['attributeValue'].strip()
				Name = data[data[data[j]['children'][1]]['children'][0]]['children'][0]
				sheet8['C{col}'.format(col=row8)] = data[Name]['attributeValue'].strip()
				Share_Allocations_Data[2] = data[Name]['attributeValue'].strip()
				sheet8['A{col}'.format(col=row8)] = Type
				sheet8['D{col}'.format(col=row8)] = Total_Shares
				Share_Allocations_Data[0] = Type
				Share_Allocations_Data[3] = Total_Shares
				row8+=1
				with open(File_path_Share_Allocations_txt,"a", encoding='utf-8')as f:
					f.write("\t".join(map(str,Share_Allocations_Data))+"\n")

	book8.save(File_path_Share_Allocations)
	book8.close()
	Dereference(sheet8)
	Dereference(book8)

def Proprietors(data,UIN):
	book9 = openpyxl.load_workbook(File_path_Proprietors)
	sheet9 = book9.active
	global row9
	collection = {}
	isCompany = 0
	for i in data:
		if 'DC_' in i:
			break
		else:
			if 'domain' in data[i] and data[i]['domain'] == 'EntityProprietor':
				isCompany = 1
				collection[i] = isCompany
			elif 'domain' in data[i] and data[i]['domain'] == 'IndividualProprietor':
				isCompany = 0
				collection[i] = isCompany
	for key,value in collection.items():
		Proprietors_Data = ['']*12
		sheet9['L{col}'.format(col=row9)] = UIN
		Proprietors_Data[11] = UIN
		if value == 1:
			Info = data[key]['children'][1]
			Details = data[Info]['children'][0]
			for k in data[Details]['children']:
				if 'attribute' in data[k] and data[k]['attribute'] == 'Name':
					sheet9['A{col}'.format(col=row9)] = data[k]['attributeValue'].strip()
					Proprietors_Data[0] = data[k]['attributeValue'].strip()
				elif 'attribute' in data[k] and data[k]['attribute'] == 'EntityNumber':
					sheet9['B{col}'.format(col=row9)] = data[k]['attributeValue'].strip() if 'attributeValue' in data[k] else ''
					Proprietors_Data[1] = data[k]['attributeValue'].strip() if 'attributeValue' in data[k] else ''
				elif 'text' in data[k] and data[k]['text']['label'] == 'Registered Office Address':
					sheet9['C{col}'.format(col=row9)] = data[data[k]['children'][0]]['text']['singleline'].strip()
					Proprietors_Data[2] = data[data[k]['children'][0]]['text']['singleline'].strip()
				elif 'text' in data[k] and data[k]['text']['label'] == 'Postal Address':
					sheet9['D{col}'.format(col=row9)] = data[data[k]['children'][0]]['text']['singleline'].strip()
					Proprietors_Data[3] = data[data[k]['children'][0]]['text']['singleline'].strip()
			Appointment = data[Info]['children'][1]
			Appointment = data[Appointment]['children'][0]
			for l in data[Appointment]['children']:
				if 'attribute' in data[l] and data[l]['attribute'] == 'StartDate':
					sheet9['E{col}'.format(col=row9)] = data[l]['attributeValue'].strip()
					Proprietors_Data[4] = data[l]['attributeValue'].strip()
			sheet9['K{col}'.format(col=row9)] = value
			Proprietors_Data[10] = value
			row9+=1
			with open(File_path_Proprietors_txt,"a", encoding='utf-8')as f:
				f.write("\t".join(map(str,Proprietors_Data))+"\n")
		else:
			sheet9['F{col}'.format(col=row9)] = data[key]['text']['singleline'].strip()
			Proprietors_Data[5] = data[key]['text']['singleline'].strip()
			Info = data[key]['children'][1]
			for j in data[Info]['children']:
				if 'text' in data[j] and data[j]['text']['label'] == 'Proprietor\'s Details':
					sheet9['G{col}'.format(col=row9)] = data[data[j]['children'][1]]['text']['dcText'].strip()
					Proprietors_Data[6] = data[data[j]['children'][1]]['text']['dcText'].strip()
				elif 'text' in data[j] and data[j]['text']['label'] == 'Residential Address':
					sheet9['H{col}'.format(col=row9)] = data[data[data[j]['children'][0]]['children'][0]]['text']['singleline'].strip()
					Proprietors_Data[7] = data[data[data[j]['children'][0]]['children'][0]]['text']['singleline'].strip()
				elif 'text' in data[j] and data[j]['text']['label'] == 'Postal Address':
					sheet9['I{col}'.format(col=row9)] = data[data[data[j]['children'][0]]['children'][0]]['text']['singleline'].strip()
					Proprietors_Data[8] = data[data[data[j]['children'][0]]['children'][0]]['text']['singleline'].strip()
				elif 'text' in data[j] and data[j]['text']['label'] == 'Additional Details':
					for k in data[data[j]['children'][0]]['children']:
						if 'attribute' in data[k] and data[k]['attribute'] == 'StartDate':
							sheet9['J{col}'.format(col=row9)] = data[k]['attributeValue'].strip()
							Proprietors_Data[9] = data[k]['attributeValue'].strip()
			sheet9['K{col}'.format(col=row9)] = value
			Proprietors_Data[10] = value
			row9+=1
			with open(File_path_Proprietors_txt,"a", encoding='utf-8')as f:
				f.write("\t".join(map(str,Proprietors_Data))+"\n")

	book9.save(File_path_Proprietors)
	book9.close()
	Dereference(sheet9)
	Dereference(book9)

def Individual_Company(Company_ID,Company_Name,UIN):
	global row2 
	global row10
	Indi_payload = json.dumps({"returnRootHtmlOnChange":"false","returnChangesOnly":"true","commands":[{"type":"conflict-check"},{"type":"view-node-button-click","id":Company_ID}]})
	Indi_headers = {"cookie":Cookie}
	try:
		Individual = requests.post(New_SearchPage_Url,data=Indi_payload,headers=Indi_headers)
		Indi_data = Individual.json()
		Indi_URL = Indi_data['redirect']
		try:
			Indi_GET = requests.get(Indi_URL)
			Indi_soup = BeautifulSoup(Indi_GET.content, 'html.parser')
			Title = Indi_soup.find('title')
			if Title.string == 'Error 500':
				global row1
				global row2
				global row3
				global row4
				global row5
				global row6
				global row7
				global row8
				global row9
				row1+=1
				row2+=1
				row3+=1
				row4+=1
				row5+=1
				row6+=1
				row7+=1
				row8+=1
				row9+=1
				return
			Indi_script = Indi_soup.find_all('script',type='text/javascript')[0].string
			Indi_View_Tree = Indi_script.split('var viewTree = ')
			Indi_View_Tree = Indi_View_Tree[1].split('for (key in viewTree)')
			Indi_View_Tree = Indi_View_Tree[0].strip()
			json_data = json.loads(Indi_View_Tree[:-1])
			Individual.close()
			Indi_GET.close()
			Dereference(Individual)
			Dereference(Indi_GET)
			Indi_soup.decompose()
			for it in json_data:
				if 'DC_' in it:
					break
				elif (it != 'root') and ('widget' in json_data[it]) and (json_data[it]['widget'] == 'wizard') and (json_data[it]['text']['shortlabel'] == 'Company Details' or json_data[it]['text']['shortlabel'] == 'Business Name'):
					Totaltabchild =json_data[it]['children']
					for currentitem in Totaltabchild:
						if(json_data[currentitem]['text']['label'] == 'General Details'):
							row2 = General_Details(it,json_data,row2,1) if json_data[it]['text']['shortlabel'] == 'Company Details' else General_Details(it,json_data,row2,0)
						elif(json_data[currentitem]['text']['label'] == 'Addresses'):
							Addresses(json_data,UIN)
						elif(json_data[currentitem]['text']['label'] == 'Directors'):
							Directors(json_data,UIN)
						elif(json_data[currentitem]['text']['label'] == 'Secretaries'):
							Secretaries(json_data,UIN)
						elif(json_data[currentitem]['text']['label'] == 'Auditors'):
							Auditors(json_data,UIN)
						elif(json_data[currentitem]['text']['label'] == 'Shareholders'):
							Shareholders(json_data,UIN)
						elif(json_data[currentitem]['text']['label'] == 'Share Allocations'):
							Share_Allocations(json_data,UIN)
						elif(json_data[currentitem]['text']['label'] == 'Proprietors'):
							Proprietors(json_data,UIN)
		except Exception as e:
			global row10
			book10 = openpyxl.load_workbook(Error_File)
			sheet10 = book10.active
			sheet10['A{col}'.format(col=row10)] = Indi_URL
			sheet10['B{col}'.format(col=row10)] = 'Not Responding'
			sheet10['C{col}'.format(col=row10)] = str(e)
			book10.save(Error_File)
			book10.close()
			Dereference(sheet10)
			Dereference(book10)
			row10+=1
	except Exception as e:
		book10 = openpyxl.load_workbook(Error_File)
		sheet10 = book10.active
		sheet10['A{col}'.format(col=row10)] = New_SearchPage_Url
		sheet10['B{col}'.format(col=row10)] = 'Not Responding'
		sheet10['C{col}'.format(col=row10)] = str(e)
		book10.save(Error_File)
		book10.close()
		Dereference(sheet10)
		Dereference(book10)
		row10+=1

def Search_Page_Info(ID_Arr,collection,rowS):
	book1 = openpyxl.load_workbook(File_path_Search_Page_Info)
	sheet1 = book1.active
	# For Name & Status
	for NS in ID_Arr:
		Search_Page_data = []
		hasUIN = 0
		UIN_test = collection[NS[1]]['children']
		for k in range(0,len(UIN_test)):
			if ('attribute' in collection[UIN_test[k]]) and (collection[UIN_test[k]]['attribute']=="businessIdentifier"):
				if ('attributeValue' in collection[UIN_test[k]]):
					hasUIN = 1
		if hasUIN == 1:
			Name_Status = collection[NS[0]]['children']
			Name = collection[Name_Status[0]]['text']['label'].strip()
			sheet1['A{col}'.format(col=rowS)] = collection[Name_Status[0]]['text']['label'].strip() if collection[Name_Status[0]]['text']['label'] else ''
			sheet1['B{col}'.format(col=rowS)] = collection[Name_Status[1]]['text']['singleline'].strip() if collection[Name_Status[1]]['text']['singleline'] else ''
			Search_Page_data.append(collection[Name_Status[0]]['text']['label'].strip())
			Search_Page_data.append(collection[Name_Status[1]]['text']['singleline'].strip())
		# For Registration No. , Company Type & Registration Date
			CRC = collection[NS[1]]['children']
			for i in range(0,len(CRC)):
				if ('attribute' in collection[CRC[i]]) and (collection[CRC[i]]['attribute']=="businessIdentifier"):
					sheet1['C{col}'.format(col=rowS)] = collection[CRC[i]]['attributeValue'].strip() if 'attributeValue' in collection[CRC[i]] else ''
					Search_Page_data.append(collection[CRC[i]]['attributeValue'].strip() if 'attributeValue' in collection[CRC[i]] else '')
					UIN_no = collection[CRC[i]]['attributeValue'].strip()
				elif ('attribute' in collection[CRC[i]]) and (collection[CRC[i]]['attribute']=="Type"):
					sheet1['D{col}'.format(col=rowS)] = collection[CRC[i]]['attributeValue'].strip() if 'attributeValue' in collection[CRC[i]] else ''
					Search_Page_data.append(collection[CRC[i]]['attributeValue'].strip() if 'attributeValue' in collection[CRC[i]] else '')
				elif ('attribute' in collection[CRC[i]]) and (collection[CRC[i]]['attribute']=="RegistrationDate"):
					sheet1['E{col}'.format(col=rowS)] = collection[CRC[i]]['attributeValue'].strip() if 'attributeValue' in collection[CRC[i]] else ''
					Search_Page_data.append(collection[CRC[i]]['attributeValue'].strip() if 'attributeValue' in collection[CRC[i]] else '')
		# For Address
			if len(NS)==3:
				if ('children' in collection[NS[2]]):
					Add = collection[NS[2]]['children']
					if ('domain' in collection[Add[0]]) and (collection[Add[0]]['domain']=="PrimaryAddress"):
						sheet1['F{col}'.format(col=rowS)] = collection[Add[0]]['text']['singleline'].strip() if collection[Add[0]]['text']['singleline'] else ''
						Search_Page_data.append(collection[Add[0]]['text']['singleline'].strip() if collection[Add[0]]['text']['singleline'] else '')
			print('Adding ' + collection[Name_Status[0]]['text']['label'].strip())
			Individual_Company(collection[Name_Status[0]]['id'],Name,UIN_no)
			rowS+=1
			with open(File_path_Search_Page_Info_txt,"a", encoding='utf-8')as f:
									f.write("\t".join(map(str,Search_Page_data))+"\n")
			with open(File_path_search_count,"a", encoding='utf-8')as fh:
				fh.write("1\n")
			# Search_Page_data = []
	book1.save(File_path_Search_Page_Info)
	Dereference(sheet1)
	Dereference(book1)
	return rowS


if __name__=='__main__':

############################################# Writing Headers for Excel Files #############################################
	File_paths= [File_path_Search_Page_Info,File_path_Addresses,File_path_Auditors,File_path_Directors,File_path_General_Details,File_path_Proprietors,File_path_Secretaries,File_path_Share_Allocations,File_path_Shareholders]
	
	directories = [
            BasePath + '\\OP',
            BasePath + '\\OPtxt',
            BasePath + '\\OPcsv',
            BasePath + '\\Error',
            BasePath + '\\Counts',
            BasePath + '\\Log'
        ]

	for directory in directories:
		if not os.path.exists(directory):
			os.makedirs(directory)


	row1 = 2
	book1 = openpyxl.Workbook()
	sheet1 = book1.active

	sheet1['A1'] = "Company name"
	sheet1['A1'].font = Font(bold=True)
	sheet1['B1'] = "Status"
	sheet1['B1'].font = Font(bold=True)
	sheet1['C1'] = "UIN"
	sheet1['C1'].font = Font(bold=True)
	sheet1['D1'] = "Company Type"
	sheet1['D1'].font = Font(bold=True)
	sheet1['E1'] = "Registration date"
	sheet1['E1'].font = Font(bold=True)
	sheet1['F1'] = "Company address"
	sheet1['F1'].font = Font(bold=True)
	book1.save(File_path_Search_Page_Info)
	Dereference(sheet1)
	Dereference(book1)

	row2 = 2
	book2 = openpyxl.Workbook()
	sheet2 = book2.active

	sheet2['A1'] = "Previous Name"
	sheet2['A1'].font = Font(bold=True)
	sheet2['B1'] = "Name Date from"
	sheet2['B1'].font = Font(bold=True)
	sheet2['C1'] = "Name Date to"
	sheet2['C1'].font = Font(bold=True)
	sheet2['D1'] = "Business Name Status"
	sheet2['D1'].font = Font(bold=True)
	sheet2['E1'] = "Registration date"
	sheet2['E1'].font = Font(bold=True)
	sheet2['F1'] = "Business Activities"
	sheet2['F1'].font = Font(bold=True)
	sheet2['G1'] = "Business Activity"
	sheet2['G1'].font = Font(bold=True)
	sheet2['H1'] = "Date of Commencement of Business Activity"
	sheet2['H1'].font = Font(bold=True)
	sheet2['I1'] = "Renewal Filing Month"
	sheet2['I1'].font = Font(bold=True)
	sheet2['J1'] = "UIN"
	sheet2['J1'].font = Font(bold=True)
	sheet2['K1'] = "Company Status"
	sheet2['K1'].font = Font(bold=True)
	sheet2['L1'] = "Status"
	sheet2['L1'].font = Font(bold=True)
	sheet2['M1'] = "Status Date From"
	sheet2['M1'].font = Font(bold=True)
	sheet2['N1'] = "Status Date to"
	sheet2['N1'].font = Font(bold=True)
	sheet2['O1'] = "Foreign Company"
	sheet2['O1'].font = Font(bold=True)
	sheet2['P1'] = "Exempt"
	sheet2['P1'].font = Font(bold=True)
	sheet2['Q1'] = "Incorporation Date"
	sheet2['Q1'].font = Font(bold=True)
	sheet2['R1'] = "Re-Registration Date"
	sheet2['R1'].font = Font(bold=True)
	sheet2['S1'] = "Have own constitution"
	sheet2['S1'].font = Font(bold=True)
	sheet2['T1'] = "Annual Return Filing Month"
	sheet2['T1'].font = Font(bold=True)
	sheet2['U1'] = "Annual Return last filed on"
	sheet2['U1'].font = Font(bold=True)
	book2.save(File_path_General_Details)
	book2.close()
	Dereference(sheet2)
	Dereference(book2)

	row3 = 2
	book3 = openpyxl.Workbook()
	sheet3 = book3.active

	sheet3['A1'] = "Registered Office Address"
	sheet3['A1'].font = Font(bold=True)
	sheet3['B1'] = "Previous Registered Office Address"
	sheet3['B1'].font = Font(bold=True)
	sheet3['C1'] = "Postal Address"
	sheet3['C1'].font = Font(bold=True)
	sheet3['D1'] = "Previous Postal Addresses"
	sheet3['D1'].font = Font(bold=True)
	sheet3['E1'] = "Principal Place of Business"
	sheet3['E1'].font = Font(bold=True)
	sheet3['F1'] = "Previous Principal Place of Business"
	sheet3['F1'].font = Font(bold=True)
	sheet3['G1'] = "UIN"
	sheet3['G1'].font = Font(bold=True)
	book3.save(File_path_Addresses)
	book3.close()
	Dereference(sheet3)
	Dereference(book3)

	row4 = 2
	book4 = openpyxl.Workbook()
	sheet4 = book4.active

	sheet4['A1'] = "Name"
	sheet4['A1'].font = Font(bold=True)
	sheet4['B1'] = "Nationality"
	sheet4['B1'].font = Font(bold=True)
	sheet4['C1'] = "Residential Address"
	sheet4['C1'].font = Font(bold=True)
	sheet4['D1'] = "Postal address"
	sheet4['D1'].font = Font(bold=True)
	sheet4['E1'] = "Appointment Date"
	sheet4['E1'].font = Font(bold=True)
	sheet4['F1'] = "Ceased Date"
	sheet4['F1'].font = Font(bold=True)
	sheet4['G1'] = "UIN"
	sheet4['G1'].font = Font(bold=True)
	book4.save(File_path_Directors)
	book4.close()
	Dereference(sheet4)
	Dereference(book4)

	row5 = 2
	book5 = openpyxl.Workbook()
	sheet5 = book5.active

	sheet5['A1'] = "Name"
	sheet5['A1'].font = Font(bold=True)
	sheet5['B1'] = "Nationality"
	sheet5['B1'].font = Font(bold=True)
	sheet5['C1'] = "Residential Address"
	sheet5['C1'].font = Font(bold=True)
	sheet5['D1'] = "Appointment Date"
	sheet5['D1'].font = Font(bold=True)
	sheet5['E1'] = "ceased Date"
	sheet5['E1'].font = Font(bold=True)
	sheet5['F1'] = "UIN"
	sheet5['F1'].font = Font(bold=True)
	book5.save(File_path_Auditors)
	book5.close()
	Dereference(sheet5)
	Dereference(book5)

	row6 = 2
	book6 = openpyxl.Workbook()
	sheet6 = book6.active

	sheet6['A1'] = "Company Name"
	sheet6['A1'].font = Font(bold=True)
	sheet6['B1'] = "UIN"
	sheet6['B1'].font = Font(bold=True)
	sheet6['C1'] = "Registered Office Address"
	sheet6['C1'].font = Font(bold=True)
	sheet6['D1'] = "Representative Name"
	sheet6['D1'].font = Font(bold=True)
	sheet6['E1'] = "Representative Postal address"
	sheet6['E1'].font = Font(bold=True)
	sheet6['F1'] = "Appointment Date"
	sheet6['F1'].font = Font(bold=True)
	sheet6['G1'] = "Name"
	sheet6['G1'].font = Font(bold=True)
	sheet6['H1'] = "Nationality"
	sheet6['H1'].font = Font(bold=True)
	sheet6['I1'] = "Residential Address"
	sheet6['I1'].font = Font(bold=True)
	sheet6['J1'] = "Postal address"
	sheet6['J1'].font = Font(bold=True)
	sheet6['K1'] = "Individual Appointment Date"
	sheet6['K1'].font = Font(bold=True)
	sheet6['L1'] = "IsCompany"
	sheet6['L1'].font = Font(bold=True)
	sheet6['M1'] = "UIN"
	sheet6['M1'].font = Font(bold=True)
	sheet6['N1'] = "Individual Ceased Date"
	sheet6['N1'].font = Font(bold=True)
	book6.save(File_path_Secretaries)
	book6.close()
	Dereference(sheet6)
	Dereference(book6)

	row7 = 2
	book7 = openpyxl.Workbook()
	sheet7 = book7.active

	sheet7['A1'] = "Company Name"
	sheet7['A1'].font = Font(bold=True)
	sheet7['B1'] = "UIN"
	sheet7['B1'].font = Font(bold=True)
	sheet7['C1'] = "Registered Office Address"
	sheet7['C1'].font = Font(bold=True)
	sheet7['D1'] = "Company Postal Address"
	sheet7['D1'].font = Font(bold=True)
	sheet7['E1'] = "Company Nominee shareholder"
	sheet7['E1'].font = Font(bold=True)
	sheet7['F1'] = "Company Appointment Date"
	sheet7['F1'].font = Font(bold=True)
	sheet7['G1'] = "Entity Name"
	sheet7['G1'].font = Font(bold=True)
	sheet7['H1'] = "Address"
	sheet7['H1'].font = Font(bold=True)
	sheet7['I1'] = "Registration Number"
	sheet7['I1'].font = Font(bold=True)
	sheet7['J1'] = "Country of Registration"
	sheet7['J1'].font = Font(bold=True)
	sheet7['K1'] = "Name of Shareholder"
	sheet7['K1'].font = Font(bold=True)
	sheet7['L1'] = "Nationality"
	sheet7['L1'].font = Font(bold=True)
	sheet7['M1'] = "Residential Address"
	sheet7['M1'].font = Font(bold=True)
	sheet7['N1'] = "Individual Postal Address"
	sheet7['N1'].font = Font(bold=True)
	sheet7['O1'] = "Individual Nominee shareholder"
	sheet7['O1'].font = Font(bold=True)
	sheet7['P1'] = "Individual Appointment Date"
	sheet7['P1'].font = Font(bold=True)
	sheet7['Q1'] = "IsCompany"
	sheet7['Q1'].font = Font(bold=True)
	sheet7['R1'] = "UIN"
	sheet7['R1'].font = Font(bold=True)
	book7.save(File_path_Shareholders)
	book7.close()
	Dereference(sheet7)
	Dereference(book7)

	row8 = 2
	book8 = openpyxl.Workbook()
	sheet8 = book8.active

	sheet8['A1'] = "Share allocation type"
	sheet8['A1'].font = Font(bold=True)
	sheet8['B1'] = "Number of Shares"
	sheet8['B1'].font = Font(bold=True)
	sheet8['C1'] = "Shareholder Name"
	sheet8['C1'].font = Font(bold=True)
	sheet8['D1'] = "Total number of shares"
	sheet8['D1'].font = Font(bold=True)
	sheet8['E1'] = "UIN"
	sheet8['E1'].font = Font(bold=True)
	book8.save(File_path_Share_Allocations)
	book8.close()
	Dereference(sheet8)
	Dereference(book8)

	row9 = 2
	book9 = openpyxl.Workbook()
	sheet9 = book9.active

	sheet9['A1'] = "Company Name"
	sheet9['A1'].font = Font(bold=True)
	sheet9['B1'] = "UIN"
	sheet9['B1'].font = Font(bold=True)
	sheet9['C1'] = "Registered Office Address"
	sheet9['C1'].font = Font(bold=True)
	sheet9['D1'] = "Postal Address"
	sheet9['D1'].font = Font(bold=True)
	sheet9['E1'] = "Company Appointment Date"
	sheet9['E1'].font = Font(bold=True)
	sheet9['F1'] = "Name"
	sheet9['F1'].font = Font(bold=True)
	sheet9['G1'] = "Nationality"
	sheet9['G1'].font = Font(bold=True)
	sheet9['H1'] = "Residential Address"
	sheet9['H1'].font = Font(bold=True)
	sheet9['I1'] = "Postal Address"
	sheet9['I1'].font = Font(bold=True)
	sheet9['J1'] = "Individual Appointment Date"
	sheet9['J1'].font = Font(bold=True)
	sheet9['K1'] = "IsCompany"
	sheet9['K1'].font = Font(bold=True)
	sheet9['L1'] = "UIN of Main Company"
	sheet9['L1'].font = Font(bold=True)
	book9.save(File_path_Proprietors)
	book9.close()
	Dereference(sheet9)
	Dereference(book9)

	row10 = 2
	book10 = openpyxl.Workbook()
	sheet10 = book10.active

	sheet10['A1'] = "URL"
	sheet10['A1'].font = Font(bold=True)
	sheet10['B1'] = "Responding Status"
	sheet10['B1'].font = Font(bold=True)
	sheet10['C1'] = "Error"
	sheet10['C1'].font = Font(bold=True)
	book10.save(Error_File)
	book10.close()
	Dereference(sheet10)
	Dereference(book10)

	Search_Page_Headers = ['Company name','Status','Registration number','Company Type','Registration date','Company address']
	General_Details_Headers = ['Previous Name','Name Date from','Name Date to','Business Name Status','Registration Date','Business Activities',
			    'Business Activity','Date of Commencement of Business Activity','Renewal Filing Month','UIN','Company Status','Status',
				'Status Date From','Status Date to','Foreign Company','Exempt','Incorporation Date','Re-Registration Date','Have own constitution','Annual Return Filing Month','Annual Return last filed on']
	Addresses_Headers = ['Registered Office Address','Previous Registered Office Address','Postal Address','Previous Postal Addresses','Principal Place of Business']
	Proprietors_Headers = ['Company Name','UIN','Registered Office Address','Postal address','Company Appointment Date','Name','Nationality','Residential Address','Postal address','Individual Appointment Date']
	Directors_Headers = ['Name','Nationality','Residential Address','Postal address','Appointment Date','ceased Date']
	Secretaries_Headers = ['Company Name','UIN','Registered Office Address','Representative Name','Representative Postal address','Appointment Date','Name','Nationality','Residential Address','Postal address','Appointment Date']
	Shareholders_Headers = ['Company Name','UIN','Registered Office Address','Company Postal Address','Company Nominee shareholder','Company Appointment Date','Entity Name','Address','Registration Number','Country of Registration',
				'Name of Shareholder','Nationality','Residential Address','Individual Postal Address','Individual Nominee shareholder','Individual Appointment Date']
	Auditors_Headers = ['Name','Nationality','Residential Address','Appointment Date','Ceased date']
	Share_Allocations_Headers = ['Share allocation type','Number of Shares','Shareholder Name','Total number of shares']

	with open(File_path_search_count,"w", encoding='utf-8')as f:
		f.write("")
	with open(File_path_Search_Page_Info_txt,"w", encoding='utf-8')as f:
		f.write("\t".join(Search_Page_Headers)+"\n")
	with open(File_path_General_Details_txt,"w", encoding='utf-8')as fw:
		fw.write("\t".join(General_Details_Headers)+"\n")
	with open(File_path_Addresses_txt,"w", encoding='utf-8')as f:
		f.write("\t".join(Addresses_Headers)+"\n")
	with open(File_path_Proprietors_txt,"w", encoding='utf-8')as fw:
		fw.write("\t".join(Proprietors_Headers)+"\n")
	with open(File_path_Directors_txt,"w", encoding='utf-8')as f:
		f.write("\t".join(Directors_Headers)+"\n")
	with open(File_path_Secretaries_txt,"w", encoding='utf-8')as fw:
		fw.write("\t".join(Secretaries_Headers)+"\n")
	with open(File_path_Shareholders_txt,"w", encoding='utf-8')as f:
		f.write("\t".join(Shareholders_Headers)+"\n")
	with open(File_path_Share_Allocations_txt,"w", encoding='utf-8')as fw:
		fw.write("\t".join(Share_Allocations_Headers)+"\n")
	with open(File_path_Auditors_txt,"w", encoding='utf-8')as f:
		f.write("\t".join(Auditors_Headers)+"\n")
	

	HomeURL = 'https://www.cipa.co.bw/'
	SearchPageURL = HomeURL + 'ng-cipa-master/ui/start/entitySearch'
	
	try:
		Driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()),options=chrome_options)
		Driver.get(SearchPageURL)
		Cookies = Driver.get_cookies()
		Cookie = Cookies[0]['name'] + '=' + Cookies[0]['value']
		New_SearchPage_Url = Driver.current_url
		
		soup = BeautifulSoup(Driver.page_source.encode(), 'html.parser')
		ID_Script = soup.find_all('script',type='text/javascript')
		Data = ID_Script[0].string
		soup.decompose()
		Driver.close()
		Driver.quit()
		View_Tree = Data.split('var viewTree = ')
		json_obj = View_Tree[1].split('for ')
		JSON_text = json_obj[0] + '\"}'
		s = json.loads(JSON_text)
		Key1 = s[s[s[s[s['root']]['children'][0]]['children'][0]]['children'][0]]['children'][1]
		# Key2 = s[s[s[s[s['root']]['children'][0]]['children'][0]]['children'][0]]['children'][2]

		for letter in alphabet[:1]:
			print('Searching for Letter : ' + letter)
			payload_1 = json.dumps({"returnRootHtmlOnChange":'false',"returnChangesOnly":'true',"commands":[{"type":"view-node-set-attribute-value","id":Key1,"value":letter}]})
			header_1 = {'content-type':'application/json','cookie':Cookie}
			try:
				response_1 = requests.post(New_SearchPage_Url,data=payload_1,headers=header_1)
				res_data_1 = response_1.json()
				data_1 = res_data_1['state']
				search_ID = ''
				for ids in data_1:
					if 'widget' in data_1[ids] and data_1[ids]['widget'] == 'search-results':
						search_ID = data_1[ids]['id']
				Dereference(response_1)
				Dereference(res_data_1)
				Dereference(data_1)
				
				Total_page = 0
				
				if letter == "a":
					payload_page = json.dumps({"returnRootHtmlOnChange":"false","returnChangesOnly":"true","commands":[{"type":"pagination-update","id":search_ID,"page":1,"size":200},{"type":"view-node-execute-rule","id":search_ID,"scope":"page-change"}]})
					header_page = {'cookie':Cookie}
					try:
						response_page = requests.post(New_SearchPage_Url,data=payload_page,headers=header_page)
						page_data = response_page.json()
						data_pg = page_data['state']
						for pg in data_pg:
							if ('widget' in data_pg[pg]) and (data_pg[pg]['widget']=='search-results'):
								Total_page = int(data_pg[pg]['kv']['ui-total'])/int(data_pg[pg]['kv']['ui-count'])
						Dereference(response_page)
						Dereference(page_data)
						Dereference(data_pg)
						Total_page = math.ceil(Total_page)+1
					except Exception as e:
							book10 = openpyxl.load_workbook(Error_File)
							sheet10 = book10.active
							sheet10['A{col}'.format(col=row10)] = New_SearchPage_Url
							sheet10['B{col}'.format(col=row10)] = 'Not Responding'
							sheet10['C{col}'.format(col=row10)] = str(e)
							book10.save(Error_File)
							book10.close()
							Dereference(sheet10)
							Dereference(book10)
							row10+=1
				else:
					for ids in data_1:
						if 'widget' in data_1[ids] and data_1[ids]['widget'] == 'search-results':
							Total_page = int(data_pg[ids]['kv']['ui-total'])/int(data_pg[ids]['kv']['ui-count'])
					Total_page = math.ceil(Total_page)+1
				for i in range(1,2):
						print('-------------------- Page : ' + str(i) + '--------------------')
						payload_2 = json.dumps({"returnRootHtmlOnChange":"false","returnChangesOnly":"true","commands":[{"type":"pagination-update","id":search_ID,"page":int(i),"size":200},{"type":"view-node-execute-rule","id":search_ID,"scope":"page-change"}]})
						header_2 = {'cookie':Cookie}
						try:
							response_2 = requests.post(New_SearchPage_Url,data=payload_2,headers=header_2)
							res_data = response_2.json()
							data = res_data['state']
							Master = []
							for item in data:
								if data[item] != None:
									if ('dos' in data[item]) and (data[item]['dos'][0]=="css-entity-search-result"):
										Master.append(data[item]['children'])
							row1 = Search_Page_Info(Master,data,row1)
							Dereference(response_2)
							Dereference(res_data)
							Dereference(data)
							book1.close()
						except Exception as e:
							book10 = openpyxl.load_workbook(Error_File)
							sheet10 = book10.active
							sheet10['A{col}'.format(col=row10)] = New_SearchPage_Url
							sheet10['B{col}'.format(col=row10)] = 'Not Responding'
							sheet10['C{col}'.format(col=row10)] = str(e)
							book10.save(Error_File)
							book10.close()
							Dereference(sheet10)
							Dereference(book10)
							row10+=1
			except Exception as e:
				book10 = openpyxl.load_workbook(Error_File)
				sheet10 = book10.active
				sheet10['A{col}'.format(col=row10)] = New_SearchPage_Url
				sheet10['B{col}'.format(col=row10)] = 'Not Responding'
				sheet10['C{col}'.format(col=row10)] = str(e)
				book10.save(Error_File)
				book10.close()
				Dereference(sheet10)
				Dereference(book10)
				row10+=1
	except Exception as e:
		book10 = openpyxl.load_workbook(Error_File)
		sheet10 = book10.active
		sheet10['A{col}'.format(col=row10)] = SearchPageURL
		sheet10['B{col}'.format(col=row10)] = 'Not Responding'
		sheet10['C{col}'.format(col=row10)] = str(e)
		book10.save(Error_File)
		book10.close()
		Dereference(sheet10)
		Dereference(book10)
		row10+=1

	print('Success')
	exit()
	
database = r"E:\ADIP Schedulers\NewWorkingDataBase\WorkingDB\InventoryDB.sqldb"

# create a database connection
conn = create_connection(database)
with conn:
	for File_path in File_paths:
		delete_task(conn, File_path)