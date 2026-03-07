import os
import sys
import traceback
import openpyxl
import re


BasePath = os.getcwd()
# BasePath= 'E:\\ADIP-PY\\OP2'
File_path_excel = BasePath + '\\InputFile\\ActivitiesComments.xlsx'
File_path_modified = BasePath + '\\OP\\ActivitiesCommentsModified.xlsx'
# File_path_excel = BasePath + '\\InputFile\\ActivitiesCommentsSample.xlsx'
# File_path_modified = BasePath + '\\OP\\ActivitiesCommentsModifiedSample.xlsx'


def exception():
    Headers_Error = ['URL', 'Not Responding', 'Error']
    error = traceback.format_exc()
    exception_type, exception_object, exception_traceback = sys.exc_info()
    print(error)


def split_activity(activity):
    try:
        hasDulplicate_flag = 0
        
        activities_list = activity.split(",")
        for lh_activity_list in activities_list:
            count = 0
            for rh_activity_list in activities_list:
                if lh_activity_list.strip() == rh_activity_list.strip():
                    count+=1
            
            if count>1:
                hasDulplicate_flag = 1
                return hasDulplicate_flag
            else:
                return hasDulplicate_flag
    
    except:
        exception()

if __name__ == '__main__':
    try:
        directories = [
            BasePath + '\\OP',
            BasePath + '\\InputFile'
        ]

        for directory in directories:
            if not os.path.exists(directory):
                os.makedirs(directory)
                
        workbook = openpyxl.load_workbook(File_path_excel)
        worksheet = workbook.active
        modified_workbook = openpyxl.Workbook()
        modified_worksheet = modified_workbook.active
        modified_worksheet.title = "Sheet1"

        modified_headers = ["idatom", "Activity Comment", "hasDulplicate"]
        modified_worksheet.append(modified_headers)
        modified_workbook.save(File_path_modified)

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            data_list = []
            idatom = row[0]
            activities = row[1]
            # hasDulplicate = 0
            
            hasDulplicate = split_activity(activities)

            data_list = [idatom,activities,hasDulplicate]
            modified_worksheet.append(data_list)
            modified_workbook.save(File_path_modified)

    except:
        exception()