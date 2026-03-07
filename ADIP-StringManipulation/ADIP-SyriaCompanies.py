import os
import sys
import traceback
import openpyxl
import re
from rapidfuzz import fuzz


BasePath = os.getcwd()
# BasePath= 'E:\\ADIP-PY\\OP2'
File_path_excel = BasePath + '\\InputFile\\BVD Kuwait - No Registers.xlsx'
# File_path_excel = BasePath + '\\InputFile\\BVD Kuwait - No RegistersSample.xlsx'
# File_path_excel = BasePath + '\\InputFile\\SyriaCompaniesInput.xlsx'
# # File_path_excel = BasePath + '\\InputFile\\SyriaCompaniesInputSample.xlsx'
File_path_modified = BasePath + '\\OP\\SyriaCompaniesInputModified.xlsx'
# File_path_modified = BasePath + '\\OP\\SyriaCompaniesInputModifiedSample95.xlsx'


def exception():
    Headers_Error = ['URL', 'Not Responding', 'Error']
    error = traceback.format_exc()
    exception_type, exception_object, exception_traceback = sys.exc_info()
    print(error)


if __name__ == '__main__':
    try:
        directories = [
            BasePath + '\\OP',
            BasePath + '\\InputFile'
        ]

        for directory in directories:
            if not os.path.exists(directory):
                os.makedirs(directory)
                
        workbook = openpyxl.load_workbook(File_path_excel)
        worksheet = workbook.active
        modified_workbook = openpyxl.Workbook()
        modified_worksheet = modified_workbook.active
        modified_worksheet.title = "Modified Names"
        
        modified_headers = ["idatom", "Name", "Town", "idatom1 Score", "Similar idatom1", "Similar name1", "idatom2 Score", "Similar idatom2", "Similar name2", "idatom3 Score", "Similar idatom3", "Similar name3"]
        modified_worksheet.append(modified_headers)
        modified_workbook.save(File_path_modified)

        exclude_list = []
        for lh_row in worksheet.iter_rows(min_row=2, values_only=True):
            lh_idatom = lh_row[0]
            lh_registerednamelocal = lh_row[1]
            lh_town = lh_row[2]
            if lh_row[0] not in exclude_list:
                data_list = []
                Similar_idatom = []
                for rh_row in worksheet.iter_rows(min_row=2, values_only=True):
                    rh_idatom = rh_row[0]
                    rh_registerednamelocal = rh_row[1]
                    rh_town = rh_row[2]
                    if lh_idatom != rh_idatom:
                        fuzzratio = round(fuzz.ratio(lh_registerednamelocal, rh_registerednamelocal), 2)
                        # if (fuzzratio>95) and (lh_town==rh_town):
                        if (fuzzratio>95):
                            exclude_list.append(rh_idatom)
                            print(f"{lh_idatom} ~ {rh_idatom} with {fuzzratio}%")
                            Similar_idatom.append(fuzzratio)
                            Similar_idatom.append(rh_idatom)
                            Similar_idatom.append(rh_registerednamelocal)
                data_list = [lh_idatom, lh_registerednamelocal, lh_town] + Similar_idatom
                modified_worksheet.append(data_list)
                modified_workbook.save(File_path_modified)

    except:
        exception()