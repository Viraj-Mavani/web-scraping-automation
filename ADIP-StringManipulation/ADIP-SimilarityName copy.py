import os
import sys
import csv
import traceback
import openpyxl
import re
from rapidfuzz import fuzz


# BasePath = os.getcwd()
BasePath= 'D:\\Projects\\CedarPython\\ADIP-StringManipulation'
# BasePath= 'E:\\ADIP-PY\\OP2'

File_path_excel = BasePath + '\\InputFile\\BVDKuwaitNewBatch.xlsx'
# File_path_excel = BasePath + '\\InputFile\\BVD Kuwait - No RegistersSample.xlsx'
File_path_modified = BasePath + '\\OP\\BVDKuwaitNewBatchModified.xlsx'
# File_path_modified = BasePath + '\\OP\\BVDKuwaitModifiedSample.xlsx'

######### CSV #########
File_path_CSV = BasePath + '\\OPcsv\\ADIP-SimilarityName_bkp.csv'
File_path_log = BasePath + '\\Log\\ADIP-SimilarityName_Log.txt'
File_path_log_Run_Flag = BasePath + '\\Log\\ADIP-SimilarityName_Run_Flag.txt'
# File_path_excel = BasePath + '\\InputFile\\SyriaCompaniesInput.xlsx'
# # File_path_excel = BasePath + '\\InputFile\\SyriaCompaniesInputSample.xlsx'
# File_path_modified = BasePath + '\\OP\\SyriaCompaniesInputModified.xlsx'
# File_path_modified = BasePath + '\\OP\\SyriaCompaniesInputModifiedSample95.xlsx'


def exception():
    Headers_Error = ['URL', 'Not Responding', 'Error']
    error = traceback.format_exc()
    exception_type, exception_object, exception_traceback = sys.exc_info()
    print(error)


def log_print(message):
    with open(File_path_log, 'a', encoding='utf-8') as file:
        file.write(message + '\n')
        file.flush()
    print(message)


if __name__ == '__main__':
    try:
        directories = [
            BasePath + '\\OP',
            BasePath + '\\OPcsv',
            BasePath + '\\Log',
            BasePath + '\\InputFile'
        ]

        for directory in directories:
            if not os.path.exists(directory):
                os.makedirs(directory)

        # First_run = True
        # if First_run:
        if not os.path.exists(File_path_log_Run_Flag):
            with open(File_path_log_Run_Flag, "a", encoding='utf-8')as f:
                f.write("")
            if os.path.exists(File_path_CSV):
                os.remove(File_path_CSV)
            if os.path.exists(File_path_log):
                os.remove(File_path_log)

        sheets = ["Part2", "Part3"]
        workbook = openpyxl.load_workbook(File_path_excel)
        
        for sheet_name in sheets:
            try:
                original_worksheet = workbook[sheet_name]
                modified_workbook = openpyxl.Workbook()
                modified_worksheet = modified_workbook.active
                modified_worksheet.title = sheet_name
                
                modified_headers = ["idatom", "Name", "Town", "idatom1 Score", "Similar idatom1", "Similar name1", "idatom2 Score", "Similar idatom2", "Similar name2", "idatom3 Score", "Similar idatom3", "Similar name3"]
                modified_worksheet.append(modified_headers)
                modified_workbook.save(File_path_modified)
                        
                with open(File_path_CSV, 'a', newline='', encoding='utf-8') as file:
                    writer = csv.writer(file)
                    if file.tell() == 0:
                        writer.writerow(modified_headers)

                # exclude_list = []
                # for lh_row in original_worksheet.iter_rows(min_row=2, values_only=True):
                #     lh_idatom = lh_row[0]
                #     lh_registerednamelocal = lh_row[1]
                #     lh_town = lh_row[2]
                #     if lh_row[0] not in exclude_list:
                #         data_list = []
                #         Similar_idatom = []
                #         for rh_row in original_worksheet.iter_rows(min_row=2, values_only=True):
                #             rh_idatom = rh_row[0]
                #             rh_registerednamelocal = rh_row[1]
                #             rh_town = rh_row[2]
                #             if lh_idatom != rh_idatom:
                #                 fuzzratio = round(fuzz.ratio(lh_registerednamelocal, rh_registerednamelocal), 2)
                #                 # if (fuzzratio>95) and (lh_town==rh_town):
                #                 if (fuzzratio>95):
                #                     exclude_list.append(rh_idatom)
                #                     log_print(f"{lh_idatom} ~ {rh_idatom} with {fuzzratio}%")
                #                     Similar_idatom.append(fuzzratio)
                #                     Similar_idatom.append(rh_idatom)
                #                     Similar_idatom.append(rh_registerednamelocal)
                #         data_list = [lh_idatom, lh_registerednamelocal, lh_town] + Similar_idatom
                #         modified_worksheet.append(data_list)
                #         modified_workbook.save(File_path_modified)
                #         with open(File_path_CSV, 'a', newline='', encoding='utf-8') as file:
                #             writer = csv.writer(file)
                #             writer.writerow(data_list)
                            
            except Exception as e:
                print(f"Error processing {sheet_name}: {e}")

    except:
        exception()