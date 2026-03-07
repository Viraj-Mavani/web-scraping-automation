import os
import sys
import traceback
import openpyxl
import re


BasePath = os.getcwd()
# BasePath= 'E:\\ADIP-PY\\OP2'
# File_path_excel = BasePath + '\\InputFile\\date_sample.xlsx'
File_path_excel = BasePath + '\\InputFile\\data.xlsx'
File_path_modified_names = BasePath + '\\OP\\arabic_names_modified.xlsx'


def exception():
    Headers_Error = ['Letter', 'URL', 'Not Responding', 'Error']
    error = traceback.format_exc()
    exception_type, exception_object, exception_traceback = sys.exc_info()
    print(error)


def replace_spaces_with_dash(name):
    try:
        words = name.split()

        for i, word in enumerate(words):
            if word in replace_space_after:
                if i < len(words) - 1:
                    words[i] = words[i] + "-" + words[i+1]
                    words.pop(i+1)
                else:
                    words.pop(i)

            elif i > 0 and words[i] in replace_space_before:
                words[i-1] = words[i-1] + "-" + words[i]
                words.pop(i)

        return " ".join(words)
    except:
        exception()  
        
def replace_dash_with_spaces(name):
    try:
        return name.replace("-", " ")
    except:
        exception()  
        

def extract_prefix(part, prefix_id_mapping, locals):
    try:
        part = replace_dash_with_spaces(part)
        
        if part in prefix_id_mapping:
            locals[0] = prefix_id_mapping[part]
            return locals
        else:
            for prefix in included_prefix_id_mapping:
                if part.startswith(prefix):
                    locals[0] = included_prefix_id_mapping[prefix]
                    locals[1] = part[len(prefix):]
                    return locals
            else:
                locals[1] = part
                return locals
    except:
        exception()

def split_arabic_name(name):
    try:
        first_name_locals = [""] * 2
        middle_1_locals = [""] * 2
        middle_2_locals = [""] * 2
        middle_3_locals = [""] * 2
        middle_4_locals = [""] * 2
        middle_5_locals = [""] * 2
        middle_6_locals = [""] * 2
        middle_7_locals = [""] * 2
        middle_8_locals = [""] * 2
        prefix_last = ""
        last_name_local = ""

        parts = name.split()

        if not last_name_local and len(parts)>=2:
            last_name_local = replace_dash_with_spaces(parts[-1])
            if parts[-2] in prefix_id_mapping:
                prefix_last = prefix_id_mapping[parts[-2]]
                parts.pop(-1)
                parts.pop(-1)
            else:
                for prefix in included_prefix_id_mapping:
                    if last_name_local.startswith(prefix) and last_name_local[len(prefix):]!='':
                        prefix_last = prefix_id_mapping[prefix]
                        last_name_local = last_name_local[len(prefix):]
                        break
                parts.pop(-1)

        for part in parts:
            if not first_name_locals[1]:
                first_name_locals = extract_prefix(part, prefix_id_mapping, first_name_locals)
            elif not middle_1_locals[1]:
                middle_1_locals = extract_prefix(part, prefix_id_mapping, middle_1_locals)
            elif not middle_2_locals[1]:
                middle_2_locals = extract_prefix(part, prefix_id_mapping, middle_2_locals)
            elif not middle_3_locals[1]:
                middle_3_locals = extract_prefix(part, prefix_id_mapping, middle_3_locals)
            elif not middle_4_locals[1]:
                middle_4_locals = extract_prefix(part, prefix_id_mapping, middle_4_locals)
            elif not middle_5_locals[1]:
                middle_5_locals = extract_prefix(part, prefix_id_mapping, middle_5_locals)
            elif not middle_6_locals[1]:
                middle_6_locals = extract_prefix(part, prefix_id_mapping, middle_6_locals)
            elif not middle_7_locals[1]:
                middle_7_locals = extract_prefix(part, prefix_id_mapping, middle_7_locals)
            elif not middle_8_locals[1]:
                middle_8_locals = extract_prefix(part, prefix_id_mapping, middle_8_locals)
        
        return first_name_locals + middle_1_locals + middle_2_locals + middle_3_locals + middle_4_locals + middle_5_locals + middle_6_locals + middle_7_locals + middle_8_locals + [prefix_last, last_name_local]
    
    except:
        exception()

if __name__ == '__main__':
    try:
        directories = [
            BasePath + '\\OP',
            BasePath + '\\InputFile'
        ]

        for directory in directories:
            if not os.path.exists(directory):
                os.makedirs(directory)
                
        workbook = openpyxl.load_workbook(File_path_excel)
        worksheet = workbook.active

        prefix_id_mapping = {
            # 'ال':'4078013',
            # 'ال':'4078014',
            'ابن':'4078015',
            'ابن':'4078016',
            'بن':'4078018',
            'بن':'4078019',
            'بنت':'4078020',
            'بنت':'4078021',
            'آل':'4078100',
            'بني':'4081604',
            'بنت':'4083456',
            'دي':'4086235',
            # 'ال':'4092242'
        }
        
        included_prefix_id_mapping = {
            # 'ال':'4078013'
        }
        
        title_id_mapping = {
            'صاحب السمو الملكي اللواء الدكتور الأمير': 3684146,
            'الكولونيل ملازم أول (متقاعد)': 3685027,
            'معالي اللواء الركن (متقاعد)': 4093095,
            'صاحبة الجلالة الإمبراطورية': 3684038,
            'صاحبة السمو الملكي الأميرة': 3684145,
            'صاحب الجلالة الإمبراطورية': 3684039,
            'سعادة الطبيب العام ملازم': 3683925,
            'صاحب السمو الملكي الأمير': 3684147,
            'الأميرال بحري (متقاعد)': 3685968,
            'سمو صاحبة السمو السيدة': 4085208,
            'صاحبة الجلالة الرسولية': 3683869,
            'العميد مهندس (متقاعد)': 3679442,
            'اللواء دكتور (متقاعد)': 3760888,
            'اللواء دكتور (متقاعد)': 3762372,
            'أدميرال بحري (متقاعد)': 3685968,
            'صاحب الجلالة الرسولية': 3683870,
            'رئيس القضاة (متقاعد)': 3681720,
            'صاحبة الجلالة الملكة': 3684052,
            'الكولونيل ملازم أول': 3685026,
            'سعادة الدكتور الشيخ': 3683923,
            'ملازم طيار (متقاعد)': 3685036,
            'جلالة الملك الشريف': 3684053,
            'سعادة العميد الركن': 4087265,
            'عميد طيار (متقاعد)': 3760887,
            'عميد طيار (متقاعد)': 3762371,
            'قائد فرقة (متقاعد)': 4083472,
            'لواء طيار (متقاعد)': 3685217,
            'السيد رئيس القضاء': 3681721,
            'سعادة اللواء طيار': 3683926,
            'الكابتن (متقاعد)': 3681489,
            'دكتور في الهندسة': 3682535,
            'كومودور (متقاعد)': 3681914,
            'العميد (متقاعد)': 3681284,
            'العميد (مهندس)': 4083140,
            'سعادة المهندسة': 4092263,
            'سمو صاحب السمو': 4085209,
            'كابتن المجموعة': 3683762,
            'الطبيب العميد': 3681279,
            'جلالة السلطان': 3684054,
            'رائد (متقاعد)': 3685212,
            'سعادة الدكتور': 3683922,
            'سعادة المهندس': 3683924,
            'عقيد (متقاعد)': 3681864,
            'قائد (متقاعد)': 3681881,
            'لواء (متقاعد)': 3685218,
            'العميد الجوي': 3680228,
            'العميد دكتور': 3681281,
            'العميد مهندس': 3681282,
            'أدميرال بحري': 3685967,
            'سعادة الشيخة': 3683920,
            'سعادة العقيد': 4083265,
            'سعادة اللواء': 4093332,
            'ضابط الطيران': 3683331,
            'قائد المنتخب': 3681487,
            'معالي الأمير': 3683927,
            'معالي الوزير': 3683921,
            'معالي الوزير': 4092798,
            'اللواء طيار': 3685216,
            'دكتور الشيخ': 3682653,
            'دكتورة شيخة': 3682654,
            'رئيس القضاء': 3681721,
            'سمو الأميرة': 4091618,
            'صاحبة السمو': 3684015,
            'قائد الجناح': 3689320,
            'معالي الشيخ': 3683928,
            'ورثة الأمير': 3683989,
            'سمو الأمير': 3684018,
            'سمو الشيخة': 3684016,
            'عميد جنرال': 3681283,
            'ملازم ثاني': 3687637,
            'ملازم طيار': 3683306,
            'ملازم طيار': 3685034,
            'مهندس طيار': 3686724,
            'ورثة الشيخ': 3683990,
            'سمو السيد': 3684017,
            'سمو الشيخ': 3684019,
            'ملازم أول': 3685029,
            'اللواء د': 3685215,
            'رقيب أول': 3687718,
            'ضابط عام': 3683547,
            'كولونيل': 4092930,
            'كومودور': 3681913,
            'الحاجة': 3683863,
            'السيدة': 3685873,
            'السيدة': 4085211,
            'العميد': 3681913,
            'القاضي': 4092241,
            'القائد': 3681880,
            'أميرال': 4092038,
            'بطريرك': 3686569,
            'السيده': 3685873,
            'الشيخة': 3687763,
            'الشيخه': 3687763,
            'الحاج': 3680143,
            'السيد': 3685865,
            'الشيخ': 3687762,
            'أستاذ': 3686965,
            'أميرة': 3686927,
            'دكتور': 3682649,
            'دكتور': 3683841,
            'سعادة': 3683919,
            'شاويش': 3687717,
            'ملازم': 3685031,
            'ملازم': 4093319,
            'مهندس': 3682940,
            'مولاي': 3685855,
            'الحج': 3683865,
            'القس': 4083003,
            'أمير': 3686925,
            'آنسة': 3685874,
            'حاجي': 3683864,
            'خوري': 4081667,
            'رئيس': 3686902,
            'شيخة': 3687763,
            'عريف': 3682052,
            'قائد': 3681913,
            'لواء': 3685214,
            'مفتي': 3685888,
            'ورثة': 3683988,
            'شيخه': 3687763,
            'أخت': 4081668,
        }

        replace_space_after = ["عبد", "ضيف","آبو","أبو","ابو","ام","أم"]
        replace_space_before = ["ألله", "الله", "الدين","بالله"]

        # prefixes = ["ال", "ابن", "بن", "بنت", "آل", "بني", "بنت", "دي", "ال"]
        data_list = []

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            idatom = row[0]
            ConcatenatedColumn = row[1]
            idtitle = None

            for title, title_id in title_id_mapping.items():
                if ConcatenatedColumn.startswith(title):
                    idtitle = title_id
                    full_name = ConcatenatedColumn[len(title):].strip()
                    break
            else:
                full_name = ConcatenatedColumn

            arabic_name = replace_spaces_with_dash(full_name)
            name_parts = split_arabic_name(arabic_name)
            data_list.append((idatom,ConcatenatedColumn,full_name,idtitle,name_parts))

        modified_workbook = openpyxl.Workbook()
        modified_worksheet = modified_workbook.active
        modified_worksheet.title = "Modified Arabic Names"

        modified_headers = ["idatom", "ConcatenatedColumn", "Full Name", "idtitle", "PrefixFirstName", "FirstName", "Prefix1", "MiddleName1", "Prefix2", "MiddleName2", "Prefix3", "MiddleName3", "Prefix4", "MiddleName4", "Prefix5", "MiddleName5", "Prefix6", "MiddleName6", "Prefix7", "MiddleName7", "Prefix8", "MiddleName8", "PrefixLastName", "LastName"]
        modified_worksheet.append(modified_headers)

        for data_tuple in data_list:
            flattened_data = [data_tuple[0]] + [data_tuple[1]] + [data_tuple[2]] + [data_tuple[3]] + [item for sublist in data_tuple[4:] for item in sublist]
            modified_worksheet.append(flattened_data)


        modified_workbook.save(File_path_modified_names)
    except:
        exception()