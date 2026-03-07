import os
import sys
import traceback
import openpyxl


BasePath = os.getcwd()
# BasePath= 'E:\\ADIP-PY\\OP2'
File_path_excel = BasePath + '\\InputFile\\date_sample.xlsx'
# File_path_excel = BasePath + '\\InputFile\\data.xlsx'
File_path_modified_names = BasePath + '\\OP\\arabic_names_modified.xlsx'


def exception():
    Headers_Error = ['Letter', 'URL', 'Not Responding', 'Error']
    error = traceback.format_exc()
    exception_type, exception_object, exception_traceback = sys.exc_info()
    print(error)


def replace_spaces_with_dash(name):
    try:
        words = name.split()

        for i, word in enumerate(words):
            if word in replace_space_after:
                if i < len(words) - 1:
                    words[i] = words[i] + "-" + words[i+1]
                    words.pop(i+1)
                else:
                    words.pop(i)

            elif i > 0 and words[i] in replace_space_before:
                words[i-1] = words[i-1] + "-" + words[i]
                words.pop(i)

        return " ".join(words)
    except:
        exception()  


if __name__=='__main__':
    try:
        # Create directories if they don't exist
        directories = [
            BasePath + '\\OP',
            BasePath + '\\InputFile'
        ]

        for directory in directories:
            if not os.path.exists(directory):
                os.makedirs(directory)
        
        # Open the Excel file
        workbook = openpyxl.load_workbook(File_path_excel)
        worksheet = workbook.active

        data_list = []
        
        # Words to replace spaces before and after
        replace_space_after = ["عبد", "ضيف"]
        replace_space_before = ["ألله", "الله", "الدين"]
        
        # Iterate through the rows in the worksheet and append (idatom, ConcatenatedColumn) tuples to the list
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            idatom = row[0]  # Assuming "idatom" is in the first column (index 0)
            arabic_name = row[1]  # Assuming "ConcatenatedColumn" is in the second column (index 1)
            print(f'Before: {arabic_name}')
            # Apply the space replacement function to the Arabic name
            arabic_name = replace_spaces_with_dash(arabic_name)
            print(f'After: {arabic_name}')

            data_list.append((idatom, arabic_name))

        # Initialize an Excel workbook and create a worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Arabic Names"

        # Write headers to the worksheet
        headers = ["idatom","ConcatenatedColumn"]
        worksheet.append(headers)


        for data_tuple in data_list:
            worksheet.append(list(data_tuple))  
        
        # Save the Excel file
        workbook.save(File_path_modified_names)
    except:
        exception()